# -*- coding: utf-8 -*-
from common import source_cn,source_en,source_link,cid_name
from db import connect_clickhouse
from os.path import abspath, join, dirname
import sys
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Fill, Border, Alignment, Protection
sys.path.insert(0, join(abspath(dirname(__file__)), '../../'))
report_path = r'{path}'.format(path=sys.path[0])

def get_top_brand(db,tbl,where):
    sql="""
        with
        case when source = 1 and shop_type < 20 and shop_type != 9 then 'tb'
             when source = 1 and shop_type > 20 or source = 1 and shop_type = 9 then 'tmall'
             when source = 2 then 'jd'
             when source = 5 then 'kaola'        
             when source = 6 then 'suning'        
             when source = 9 then 'jiuxian'
             when source = 11 then 'douyin'             
             else '其他' end as "平台",
        IF(source*100+shop_type IN [109,112,122,124,127,221,222,321,322,412,521,522,621,622,712,821,822,1121,1122], 'Cross-border', 'Domestic')  AS "跨境"
        select alias_all_bid,dictGetOrDefault('all_brand', 'name', tuple(toUInt32(alias_all_bid)), '') AS "品牌",
        sum(sales)/100 AS sales_value
        from {tbl}
        {where}
        group by alias_all_bid
        order by sales_value desc 
        limit 10  
        """.format(tbl=tbl,where=where)
    ret = get_data(db,sql,as_dict=False)
    ret = [r[0] for r in ret]
    print(ret)
    return ret

def get_Players_report(db, tbl,brand_where, players_where,total_type):
    brand_in = get_top_brand(db, tbl, brand_where)
    if total_type != 'total':
        sql = '''
                with
                case when source = 1 and shop_type < 20 and shop_type != 9 then 'tb'
                     when source = 1 and shop_type > 20 or source = 1 and shop_type = 9 then 'tmall'
                     when source = 2 then 'jd'
                     when source = 5 then 'kaola'        
                     when source = 6 then 'suning'        
                     when source = 9 then 'jiuxian'
                     when source = 11 then 'douyin'             
                     else '其他' end as "平台",
                IF(source*100+shop_type IN [109,112,122,124,127,221,222,321,322,412,521,522,621,622,712,821,822,1121,1122], 'Cross-border', 'Domestic')  AS "跨境"
                select 
                transform(alias_all_bid,{brand_in}, [1,2,3,4,5,6,7,8,9,10],999) rank,
                case when alias_all_bid  in {brand_in} then alias_all_bid
                    else 0 end as alias_all_bid
                ,dictGetOrDefault('all_brand', 'name', tuple(toUInt32(alias_all_bid)), 'Others')"品牌"
                ,toStartOfMonth(pkey) AS Gmonth,
                           sum(sales)/100 AS columns1,
                           sum(num) AS columns2, 
                           sum(sales)/100/sum(num) as columns3,
                           sum (num* toFloat64OrZero(`spSKU件数`)) columns4
                from {tbl}
                {where}
                group by alias_all_bid,rank,Gmonth
            '''.format(tbl=tbl,brand_in=brand_in,where=players_where)
        ret = get_data(db, sql)
        df = pd.DataFrame(ret)
        df.loc[len(df)] = {'rank': 999, 'alias_all_bid': 0, "品牌": 'Others', 'Gmonth': datetime(2024, 12, 1),'columns1': '-', 'columns2': '-', 'columns3': '-', 'columns4': '-'}

    else:
        sql = '''
            with
            case when source = 1 and shop_type < 20 and shop_type != 9 then 'tb'
                 when source = 1 and shop_type > 20 or source = 1 and shop_type = 9 then 'tmall'
                 when source = 2 then 'jd'
                 when source = 5 then 'kaola'        
                 when source = 6 then 'suning'        
                 when source = 9 then 'jiuxian'
                 when source = 11 then 'douyin'             
                 else '其他' end as "平台",
            IF(source*100+shop_type IN [109,112,122,124,127,221,222,321,322,412,521,522,621,622,712,821,822,1121,1122], 'Cross-border', 'Domestic')  AS "跨境"
            select 
            999 rank
            ,0 as alias_all_bid
            ,'Online Total' as "品牌"
            ,toStartOfMonth(pkey) AS Gmonth,
                       sum(sales)/100 AS columns1,
                       sum(num) AS columns2, 
                       sum(sales)/100/sum(num) as columns3,
                       sum (num* toFloat64OrZero(`spSKU件数`)) columns4
            from {tbl}
            {where}
            group by alias_all_bid,rank,Gmonth
        '''.format(tbl=tbl, brand_in=brand_in, where=players_where)
        ret = get_data(db,sql)
        df = pd.DataFrame(ret)
        df.loc[len(df)] = {'rank':999,'alias_all_bid':0,"品牌":'Online Total','Gmonth':datetime(2024, 12, 1),'columns1':'-','columns2':'-','columns3':'-','columns4':'-'}
    return df

def get_Format_report(db, tbl,brand_where, players_where,total_type):
    if total_type != 'total':
        sql = '''
                with
                case when source = 1 and shop_type < 20 and shop_type != 9 then 'tb'
                     when source = 1 and shop_type > 20 or source = 1 and shop_type = 9 then 'tmall'
                     when source = 2 then 'jd'
                     when source = 5 then 'kaola'        
                     when source = 6 then 'suning'        
                     when source = 9 then 'jiuxian'
                     when source = 11 then 'douyin'             
                     else '其他' end as "平台",
                IF(source*100+shop_type IN [109,112,122,124,127,221,222,321,322,412,521,522,621,622,712,821,822,1121,1122], 'Cross-border', 'Domestic')  AS "跨境",
                case when "sp剂型"= '片剂' then 'tablet'
                     when "sp剂型"= '液体' then 'ready-to-drink'
                     when "sp剂型"= '丸剂' then 'pill'
                     when "sp剂型"= '软糖' then 'gummy'
                     when "sp剂型"= '胶囊' then 'capsule'
                     when "sp剂型"= '粉剂' then 'powder'
                     when "sp剂型"= '果冻' then 'jelly'
                     when "sp剂型"= '泡腾片' then 'effervescent tablets'     
                     else 'others' end as "剂型"
                select 
                "剂型"
                ,transform("剂型",['capsule','effervescent tablets','gummy','jelly','pill','powder','ready-to-drink','tablet','others'], [1,2,3,4,5,6,7,8,9],999) rank 
                ,toStartOfMonth(pkey) AS Gmonth,
                           sum(sales)/100 AS columns1,
                           sum(num) AS columns2, 
                           sum(sales)/100/sum(num) as columns3,
                           sum (num* toFloat64OrZero(`spSKU件数`)) columns4
                from {tbl}
                {where}
                group by "剂型",Gmonth
            '''.format(tbl=tbl, where=players_where)

        ret = get_data(db,sql)
        df = pd.DataFrame(ret)
        df.loc[len(df)] = {'rank':1,"剂型":'capsule','Gmonth':datetime(2024, 12, 1),'columns1':'-','columns2':'-','columns3':'-','columns4':'-'}
        df.loc[len(df)] = {'rank':2,"剂型":'effervescent tablets','Gmonth':datetime(2024, 12, 1),'columns1':'-','columns2':'-','columns3':'-','columns4':'-'}
        df.loc[len(df)] = {'rank':3,"剂型":'gummy','Gmonth':datetime(2024, 12, 1),'columns1':'-','columns2':'-','columns3':'-','columns4':'-'}
        df.loc[len(df)] = {'rank':4,"剂型":'jelly','Gmonth':datetime(2024, 12, 1),'columns1':'-','columns2':'-','columns3':'-','columns4':'-'}
        df.loc[len(df)] = {'rank':5,"剂型":'pill','Gmonth':datetime(2024, 12, 1),'columns1':'-','columns2':'-','columns3':'-','columns4':'-'}
        df.loc[len(df)] = {'rank':6,"剂型":'powder','Gmonth':datetime(2024, 12, 1),'columns1':'-','columns2':'-','columns3':'-','columns4':'-'}
        df.loc[len(df)] = {'rank':7,"剂型":'ready-to-drink','Gmonth':datetime(2024, 12, 1),'columns1':'-','columns2':'-','columns3':'-','columns4':'-'}
        df.loc[len(df)] = {'rank':8,"剂型":'tablet','Gmonth':datetime(2024, 12, 1),'columns1':'-','columns2':'-','columns3':'-','columns4':'-'}
        df.loc[len(df)] = {'rank':9,"剂型":'others','Gmonth':datetime(2024, 12, 1),'columns1':'-','columns2':'-','columns3':'-','columns4':'-'}
    else:
        sql = '''
            with
            case when source = 1 and shop_type < 20 and shop_type != 9 then 'tb'
                 when source = 1 and shop_type > 20 or source = 1 and shop_type = 9 then 'tmall'
                 when source = 2 then 'jd'
                 when source = 5 then 'kaola'        
                 when source = 6 then 'suning'        
                 when source = 9 then 'jiuxian'
                 when source = 11 then 'douyin'             
                 else '其他' end as "平台",
            IF(source*100+shop_type IN [109,112,122,124,127,221,222,321,322,412,521,522,621,622,712,821,822,1121,1122], 'Cross-border', 'Domestic')  AS "跨境",
            'Online Total' as "剂型"
            select 
            "剂型"
            ,transform("剂型",['capsule','effervescent tablets','gummy','jelly','pill','powder','ready-to-drink','tablet','others'], [1,2,3,4,5,6,7,8,9],999) rank 
            ,toStartOfMonth(pkey) AS Gmonth,
                       sum(sales)/100 AS columns1,
                       sum(num) AS columns2, 
                       sum(sales)/100/sum(num) as columns3,
                       sum (num* toFloat64OrZero(`spSKU件数`)) columns4
            from {tbl}
            {where}
            group by "剂型",Gmonth
        '''.format(tbl=tbl, where=players_where)
        ret = get_data(db, sql)
        df = pd.DataFrame(ret)
        df.loc[len(df)] = {'rank': 999, "剂型": 'Online Total', 'Gmonth': datetime(2024, 12, 1), 'columns1': '-','columns2': '-', 'columns3': '-', 'columns4': '-'}
    return df

def get_data(db,sql,as_dict=True):

    data = db.query_all(sql,as_dict=as_dict)
    # print(data)
    return data

def get_pivot_df(df,index, columns, values, aggfunc, fill_value):

    # 创建透视表
    pivot_df = df.pivot_table(
        index=index,
        columns=columns,
        values=values,
        aggfunc=aggfunc,
        fill_value=fill_value
    )

    pivot_df.reset_index(inplace=True)
    # print(pivot_df.shape)
    pivot_df.dropna(how='all', inplace=True)
    return pivot_df

def write_xlsx(data,row_s,col_s,work_book,sheet_name):
    sheet = work_book[sheet_name]
    max_row,max_col = data.shape
    data = np.array(data).tolist()
    data = [d[1:] for d in data]
    for r in range(max_row):
        for c in range(max_col-1):
            sheet.cell(row_s+r,col_s+c).value = data[r][c]
    return work_book,max_row+row_s

def run_Players_report(Players_dict,tbl,work_book):
    sheet_name = 'Key Players Performance'  # 报告要保存在哪一个sheet
    db = connect_clickhouse('chsop')

    row_s, cols = 4, 1
    for source in Players_dict:
        Player_dict = Players_dict[source]
        for players, where in Player_dict.items():
            df = get_Players_report(db, tbl, where['brand_where'], where['players_where'],total_type=source)
            index = ['rank', '品牌']
            columns = 'Gmonth'
            values = ['columns1', 'columns2', 'columns3', 'columns4']
            aggfunc = {
                'columns1': 'sum',
                'columns2': 'sum',
                'columns3': 'sum',
                'columns4': 'sum'
            }
            pivot_df = get_pivot_df(df, index=index, columns=columns, values=values, aggfunc=aggfunc,
                                    fill_value=0)
            work_book, row_s = write_xlsx(pivot_df, row_s, 1, work_book, sheet_name)
            row_s += 1
        row_s += 1
    return work_book
def run_Format_report(Players_dict,tbl,work_book):
    sheet_name = 'Key Format Performance'  # 报告要保存在哪一个sheet
    db = connect_clickhouse('chsop')
    row_s, cols = 4, 1
    for source in Players_dict:
        Player_dict = Players_dict[source]
        for players, where in Player_dict.items():
            df = get_Format_report(db, tbl, where['brand_where'], where['players_where'],total_type=source)
            index = ['rank','剂型']
            columns = 'Gmonth'
            values = ['columns1', 'columns2', 'columns3', 'columns4']
            aggfunc = {
                'columns1': 'sum',
                'columns2': 'sum',
                'columns3': 'sum',
                'columns4': 'sum'
            }
            pivot_df = get_pivot_df(df, index=index, columns=columns, values=values, aggfunc=aggfunc,
                                    fill_value=0)
            print(pivot_df)
            work_book, row_s = write_xlsx(pivot_df, row_s, 1, work_book, sheet_name)
            row_s += 1
        row_s += 1
    return work_book
def get_Players_dict(brand_where,players_where):
    Players_dict = {
        'tb':{
            'Taobao': {
                'brand_where': brand_where + '''
                    and "平台"='tb'
                         ''',
                'players_where':players_where + '''
                and "平台"='tb'
                     '''
            },
            'Taobao&Domestic': {
                'brand_where': brand_where + '''
                    and "平台"='tb'
                    and "跨境"='Domestic'
                             ''',
                'players_where': players_where + '''
                    and "平台"='tb'
                    and "跨境"='Domestic'
                         '''
            },
            'Taobao&Cross-border': {
                'brand_where': brand_where + '''
                        and "平台"='tb'
                        and "跨境"='Cross-border'
                                 ''',
                'players_where': players_where + '''
                        and "平台"='tb'
                        and "跨境"='Cross-border'
                             '''
            }
        },
        'tmall':{
            'Tmall': {
                'brand_where': brand_where + '''
                and "平台"='tmall'
                     ''',
                'players_where': players_where + '''
            and "平台"='tmall'
                 '''
            },
            'Tmall&Domestic': {
                'brand_where': brand_where + '''
                and "平台"='tmall'
                and "跨境"='Domestic'
                         ''',
                'players_where': players_where + '''
                and "平台"='tmall'
                and "跨境"='Domestic'
                     '''
            },
            'Tmall&Cross-border': {
                'brand_where': brand_where + '''
                    and "平台"='tmall'
                    and "跨境"='Cross-border'
                             ''',
                'players_where': players_where + '''
                    and "平台"='tmall'
                    and "跨境"='Cross-border'
                         '''
            }
        },
        'JD':{
            'JD': {
                'brand_where': brand_where + '''
                and "平台"='jd'
                     ''',
                'players_where': players_where + '''
            and "平台"='jd'
                 '''
            },
            'JD&Domestic': {
                'brand_where': brand_where + '''
                and "平台"='jd'
                and "跨境"='Domestic'
                         ''',
                'players_where': players_where + '''
                and "平台"='jd'
                and "跨境"='Domestic'
                     '''
            },
            'JD&Cross-border': {
                'brand_where': brand_where + '''
                    and "平台"='jd'
                    and "跨境"='Cross-border'
                             ''',
                'players_where': players_where + '''
                    and "平台"='jd'
                    and "跨境"='Cross-border'
                         '''
            }
        },
        'Douyin':{
            'Douyin': {
                'brand_where': brand_where + '''
                and "平台"='douyin'
                     ''',
                'players_where': players_where + '''
            and "平台"='douyin'
                 '''
            },
            'Douyin&Domestic': {
                'brand_where': brand_where + '''
                and "平台"='douyin'
                and "跨境"='Domestic'
                         ''',
                'players_where': players_where + '''
                and "平台"='douyin'
                and "跨境"='Domestic'
                     '''
            },
            'Douyin&Cross-border': {
                'brand_where': brand_where + '''
                    and "平台"='douyin'
                    and "跨境"='Cross-border'
                             ''',
                'players_where': players_where + '''
                    and "平台"='douyin'
                    and "跨境"='Cross-border'
                         '''
            }
        }
    }
    return Players_dict
def main():
    #######################################  取报告可变的全局参数，可以任意改变  #############################################
    template = 'Eye Health Supplement in China EC Market - sample.xlsx'  # 报告魔板工作簿
    work_book = load_workbook(r'..\report\92312\\' + template)
    output = '91130 output_test.xlsx' #报告最终输出的工作簿名称
    tbl = 'sop_e.entity_prod_91130_E'  #从哪一张e表取报告
    Players_brand_where = '''
            where pkey>='2024-01-01' and pkey<'2024-12-01'
            and alias_all_bid not in (0,26,4023)
            and "sp子品类"='植物精华' 
            and or ("sp种类"='叶黄素',"sp复合种类" in ('蓝莓叶黄素','越橘叶黄素胡萝卜素','叶黄素越橘','叶黄素越橘蓝莓','蓝莓黑加仑叶黄素','野樱莓叶黄素','蓝莓叶黄素β-胡萝卜','虾青素叶黄素'))
        '''
    Players_players_where = '''
            where pkey>='2022-01-01' and pkey<'2025-01-01'
            and "sp子品类"='植物精华' 
            and or ("sp种类"='叶黄素',"sp复合种类" in ('蓝莓叶黄素','越橘叶黄素胡萝卜素','叶黄素越橘','叶黄素越橘蓝莓','蓝莓黑加仑叶黄素','野樱莓叶黄素','蓝莓叶黄素β-胡萝卜','虾青素叶黄素'))
        '''

    Format_brand_where = '''
                where pkey>='2024-01-01' and pkey<'2024-12-01'
                and alias_all_bid not in (0,26,4023)
                and "sp子品类"='植物精华' 
                and or ("sp种类"='叶黄素',"sp复合种类" in ('蓝莓叶黄素','越橘叶黄素胡萝卜素','叶黄素越橘','叶黄素越橘蓝莓','蓝莓黑加仑叶黄素','野樱莓叶黄素','蓝莓叶黄素β-胡萝卜','虾青素叶黄素'))
            '''
    Format_players_where = '''
                where pkey>='2022-01-01' and pkey<'2025-01-01'
                and "sp子品类"='植物精华' 
                and or ("sp种类"='叶黄素',"sp复合种类" in ('蓝莓叶黄素','越橘叶黄素胡萝卜素','叶黄素越橘','叶黄素越橘蓝莓','蓝莓黑加仑叶黄素','野樱莓叶黄素','蓝莓叶黄素β-胡萝卜','虾青素叶黄素'))
            '''
    Players_dict = get_Players_dict(Players_brand_where,Players_players_where)
    Format_dict = get_Players_dict(Format_brand_where, Format_players_where)
    total_dict = {
        'total': {
            'total': {
                'brand_where': Format_brand_where + ''' 
                            and "平台" in ('tb','tmall','jd','douyin')
                                 ''',
                'players_where': Format_players_where + ''' 
                        and "平台" in ('tb','tmall','jd','douyin')
                             '''
            },
        },
        'total_detaile': {
            'total_detaile': {
                'brand_where': Format_brand_where + ''' 
                                and "平台" in ('tb','tmall','jd','douyin')
                                     ''',
                'players_where': Format_players_where + ''' 
                            and "平台" in ('tb','tmall','jd','douyin')
                                 '''
            },
        }
    }
    Players_dict = {**total_dict, **Players_dict}
    Format_dict = {**total_dict, **Format_dict}
    work_book = run_Players_report(Players_dict,tbl,work_book)
    work_book = run_Format_report(Format_dict, tbl, work_book)
    work_book.save(r'..\report\92312\\' + output)

if __name__ == '__main__':
    main()
