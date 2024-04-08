# -*- coding: utf-8 -*-
from common import source_cn,source_en,source_link,cid_name
from db import connect_clickhouse
from os.path import abspath, join, dirname
import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook
sys.path.insert(0, join(abspath(dirname(__file__)), '../../'))
report_path = r'{path}'.format(path=sys.path[0])

def get_topbrand_info(where,tbl,limit=10):

    sql="""
        SELECT alias_all_bid,dictGetOrDefault('all_brand', 'name', tuple(toUInt32(alias_all_bid)), '') AS "品牌名",
        SUM(num)"销量",SUM(sales)/100000"销售额" FROM {tbl}
        {where}
        GROUP BY alias_all_bid
        ORDER BY "销售额" DESC
        LIMIT {limit};
        """.format(tbl=tbl,where=where,limit=limit)

    return sql

def get_topitem_info(where,tbl):

    sql="""
        SELECT item_id,argMax(name,num) "宝贝名称",
        SUM(num)"销量",SUM(sales)/100000"销售额" FROM {tbl}
        {where}
        GROUP BY item_id
        ORDER BY "销售额" DESC
        LIMIT 10;
        """.format(tbl=tbl,where=where)

    return sql

def get_data(db,sql):

    data = db.query_all(sql,as_dict=True)

    return data

def get_topbrand_report(db,tbl,category_list):

    for i,category in enumerate(category_list):
        where = f"""where pkey>='2023-01-01' and pkey<'2024-01-01' and "sp子品类"='{category}'"""
        sql = get_topbrand_info(where=where,tbl=tbl)
        data = get_data(db,sql)
        print(data)
        brand_sales = [[data[i].get("品牌名", ''), data[i].get("销售额", '')] if i < len(data) else ['',''] for i in range(0, 10)]
        data = [[category+" Top10品牌",'FY2023']] + brand_sales
        if i == 0:
            rr = data
        else:
            rr += data
    rr = pd.DataFrame(rr,columns=['Sales Value(000 RMB)',''])
    return rr

def get_topitem_report(db,tbl,category_list):

    where = """where pkey>='2023-01-01' and pkey<'2024-01-01' and "sp子品类" in {category_list} """.format(category_list=category_list)
    sql = get_topbrand_info(where=where, tbl=tbl, limit=5)
    data = get_data(db, sql)
    for n, d in enumerate(data):
        for i, category in enumerate(category_list):
            sql = get_topitem_info(where="""where pkey>='2023-01-01' and pkey<'2024-01-01' and "sp子品类"='{category}' and alias_all_bid={alias_all_bid} """.format(category=category, alias_all_bid=d.get("alias_all_bid", '')), tbl=tbl)
            data = get_data(db, sql)
            item_sales = [[data[k].get("item_id", ''), data[k].get("宝贝名称", ''), data[k].get("销售额", '')] if k < len(data) else ['','',''] for k in range(0, 10)]
            data = [[category, '', '']] + item_sales
            if i == 0:
                rr0 = [['指定品牌 Top10宝贝', '宝贝名称', 'FY2023']] + data
            else:
                rr0 += data
        if n == 0:
            rr = [rr0[j] + [''] for j in range(0, len(rr0))]
            columns = ['Sales Value(000 RMB)', '宝贝名称', '', '']
        else:
            columns += ['Sales Value(000 RMB)', '宝贝名称', '', '']
            rr = [rr[j] + rr0[j] + [''] for j in range(0, len(rr0))]
    rr = pd.DataFrame(rr, columns=columns)
    return rr

def write_xlsx(data,row_s,col_s,work_book,sheet_name):
    sheet = work_book[sheet_name]
    max_row,max_col = data.shape
    data = np.array(data).tolist()
    for r in range(max_row):
        for c in range(max_col):
            sheet.cell(row_s+r,col_s+c).value = data[r][c]
    return work_book

def main():
    tbl = 'sop_e.entity_prod_91559_E_bake'
    template = '安利报价_头发护理_20240306 template.xlsx'
    output = '安利报价_头发护理_20240306 Output.xlsx'
    sheet_name = 'Template'
    db = connect_clickhouse('chsop')
    work_book = load_workbook(r'..\report\91559\\' + template)
    category_list = ['洗发水','套装','头发护理','护发素','头皮护理','干性洗发','头皮清洁']
    # 写入top_brand
    rr = get_topbrand_report(db=db, tbl=tbl, category_list=category_list)
    work_book = write_xlsx(rr, 3, 2, work_book, sheet_name)
    # 写入top_item
    rr = get_topitem_report(db=db,tbl=tbl,category_list=category_list)
    work_book = write_xlsx(rr, 3, 5, work_book, sheet_name)
    # 保存工作簿
    work_book.save(r'..\report\91559\\' + output)


if __name__ == '__main__':
    main()