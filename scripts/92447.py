# -*- coding: utf-8 -*-
from common import source_cn,source_en,source_link,cid_name
from db import connect_clickhouse
from os.path import abspath, join, dirname
import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Fill, Border, Alignment, Protection
sys.path.insert(0, join(abspath(dirname(__file__)), '../../'))
report_path = r'{path}'.format(path=sys.path[0])

def get_brand_info(where,tbl):
    sql="""
        SELECT
            -- alias_all_bid,
            BRAND_NAME,
            -- sales_value,
            -- sales_volume,
            RANK() OVER (
            ORDER BY sales_volume DESC) AS rank_by_sales_volume,
            RANK() OVER (
            ORDER BY sales_value DESC) AS rank_by_sales_value
        FROM(
            SELECT
                alias_all_bid,
                dictGetOrDefault('all_brand',
                'name',
                tuple(toUInt32(alias_all_bid)),
                '') AS BRAND_NAME,
                sum(num * sign) AS sales_volume,
                sum(sales * sign / 100) AS sales_value
            FROM
                {tbl}
            {where}
            GROUP BY
                alias_all_bid
        ) AS subquery
        ORDER BY
            sales_value DESC
        limit 150
        """.format(tbl=tbl,where=where)

    return sql

def get_cate10_info(where,tbl):
    sql="""
        SELECT res.BRAND_NAME,res.sales_volume,res.sales_value,brand.r FROM (
            WITH brand_dict AS (
                SELECT [22211,50647,68110,84613,24115,84516,989603,85792,143768,265393,51284,3626,50893,6786109,2833661,50860,5708533,2029107,720056,2101482,9488,6900990,10003,92794,1458202,143827,90748,84721,6963861,1933286,4495107,13630,51635,748250,6391699,2413988,2030094,1792664,760178,411118,6269251,2321041,85671,172,5904176,21409,220335,2638245,1190992,90785,69375,2413984,50729,4551931,51043,6208071,68552,147131,2529353,1345681,3224123,6314484,3242499,4379403,5905580,68128,3163182,653560,4820820,50726,26415,6522071,6510204,3638420,6581968,6894982,6764068,219747,250115,5175840,6729527,1039719,5392220,2695780,5075949,51512,6760922,5918699,3718928,657838,262877,2362554,540637,5787445,4411073,3227263,377854,6299865,84656,3852,3015947,1456927,6065371,68133,5194935,749508,5075962,6134788,2349967,3699024,6377258,91193,48476,91703,7009129,4886954,271885,1407699,5762449,747161,4254756,613,91858,143792,220309,3628811,68142,6040434,401474,2862,189679,7320688,6062982,3499163,2793749,35174,4421253,157620,877003,6054982,6090730,51694,51054,6304548,30649,8980,7034574,3587615,5281611,1265639,68462,4665225,90755,3675332,3671491,5279003,3935893,220620,6433403,1022073,6126122,7559176,332858,6286535,6565777,57796,7021834,3487758,3746449,3643832,7007875,6774148,3586605,3746594,3713969,3675338,6455456,893188,5912550,5181271,92190,219562,1993817,6008849,220358,6772675,7106604,1497497,50712,6076997,93535,90777,1792801,3770262,4748146,6421811,2827306,32563,1215550,5749514,219829,42464,5973222,143951,1907440,6654142,7194519,91760] bids,[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,63,64,65,66,67,68,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101,102,103,104,105,106,107,108,109,110,111,112,113,114,115,116,117,118,119,120,121,122,123,124,125,126,127,128,129,130,131,132,133,134,135,136,137,138,139,140,141,142,143,144,145,146,147,148,149,150,151,152,153,154,155,156,157,158,159,160,161,162,163,164,165,166,167,168,169,170,171,172,173,174,175,176,177,178,179,180,181,182,183,184,185,186,187,188,189,190,191,192,193,194,195,196,197,198,199,200,201,202,203,204,205,206,207,208] rs)
            SELECT 
                bid,
                r
            FROM brand_dict
            ARRAY JOIN 
                brand_dict.bids AS bid,
                brand_dict.rs AS r
            ) brand
        LEFT JOIN (
            SELECT
                alias_all_bid,
                dictGetOrDefault('all_brand',
                'name',
                tuple(toUInt32(alias_all_bid)),
                '') AS BRAND_NAME,
                sum(num * sign) AS sales_volume,
                sum(sales * sign / 100) AS sales_value,
                transform(alias_all_bid,[22211,50647,68110,84613,24115,84516,989603,85792,143768,265393,51284,3626,50893,6786109,2833661,50860,5708533,2029107,720056,2101482,9488,6900990,10003,92794,1458202,143827,90748,84721,6963861,1933286,4495107,13630,51635,748250,6391699,2413988,2030094,1792664,760178,411118,6269251,2321041,85671,172,5904176,21409,220335,2638245,1190992,90785,69375,2413984,50729,4551931,51043,6208071,68552,147131,2529353,1345681,3224123,6314484,3242499,4379403,5905580,68128,3163182,653560,4820820,50726,26415,6522071,6510204,3638420,6581968,6894982,6764068,219747,250115,5175840,6729527,1039719,5392220,2695780,5075949,51512,6760922,5918699,3718928,657838,262877,2362554,540637,5787445,4411073,3227263,377854,6299865,84656,3852,3015947,1456927,6065371,68133,5194935,749508,5075962,6134788,2349967,3699024,6377258,91193,48476,91703,7009129,4886954,271885,1407699,5762449,747161,4254756,613,91858,143792,220309,3628811,68142,6040434,401474,2862,189679,7320688,6062982,3499163,2793749,35174,4421253,157620,877003,6054982,6090730,51694,51054,6304548,30649,8980,7034574,3587615,5281611,1265639,68462,4665225,90755,3675332,3671491,5279003,3935893,220620,6433403,1022073,6126122,7559176,332858,6286535,6565777,57796,7021834,3487758,3746449,3643832,7007875,6774148,3586605,3746594,3713969,3675338,6455456,893188,5912550,5181271,92190,219562,1993817,6008849,220358,6772675,7106604,1497497,50712,6076997,93535,90777,1792801,3770262,4748146,6421811,2827306,32563,1215550,5749514,219829,42464,5973222,143951,1907440,6654142,7194519,91760],[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,63,64,65,66,67,68,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101,102,103,104,105,106,107,108,109,110,111,112,113,114,115,116,117,118,119,120,121,122,123,124,125,126,127,128,129,130,131,132,133,134,135,136,137,138,139,140,141,142,143,144,145,146,147,148,149,150,151,152,153,154,155,156,157,158,159,160,161,162,163,164,165,166,167,168,169,170,171,172,173,174,175,176,177,178,179,180,181,182,183,184,185,186,187,188,189,190,191,192,193,194,195,196,197,198,199,200,201,202,203,204,205,206,207,208],0) rank
            FROM
                {tbl}
            {where}
            GROUP BY
                alias_all_bid
            ) res
        ON brand.bid = res.alias_all_bid
        ORDER BY brand.r
        """.format(tbl=tbl,where=where)

    return sql

def get_data(db,sql,as_dict=True):

    data = db.query_all(sql,as_dict=as_dict)

    return data

def get_cate_dict():
    cate_dict={
    "羊毛被/驼毛被":[15256,26252,34718,126048001],
    "蚕丝被":[15255,26241,34714,126050001],
    "被套":[15275,20987,34677,50001865,50008263],
    "羽绒/羽毛被":[15254,26172,34719,126046001],
    "休闲毯/毛毯/绒毯":[15273,17393,21160,26980,34720,34744,34745,50001871],
    "棉花被":[15258,26240,34715,126054001],
    "化纤被":[26239,34717,126052001,15257],
    "枕头/枕芯":[15259,15260,15261,15262,15263,15264,21104,26095,26161,26236,26237,27306,27324,34704,34705,34706,34707,34708,34709,34710,34711,34712,37712,37713,37714,37715,37716,37717,37718,37719,50002777,201304409,201309307,201309312,202146523,202149818],
    "床品套件/四件套/多件套":[15251,15252,15253,21273,34684,34685,34686,34687,290903,50008779,201305913,201310106,201663213],
    "床垫":[15274,21041,34695,34696,34697,34698,34699,34700,34701,34702,50008565,50019373]
   }

    return cate_dict

def get_cate_dict2():
    cate_dict={
    "床上用品":[0],
    "10类目":[15256,26252,34718,126048001,15255,26241,34714,126050001,15275,20987,34677,50001865,50008263,15254,26172,34719,126046001,15273,17393,21160,26980,34720,34744,34745,50001871,15258,26240,34715,126054001,26239,34717,126052001,15257,15259,15260,15261,15262,15263,15264,21104,26095,26161,26236,26237,27306,27324,34704,34705,34706,34707,34708,34709,34710,34711,34712,37712,37713,37714,37715,37716,37717,37718,37719,50002777,201304409,201309307,201309312,202146523,202149818,15251,15252,15253,21273,34684,34685,34686,34687,290903,50008779,201305913,201310106,201663213,15274,21041,34695,34696,34697,34698,34699,34700,34701,34702,50008565,50019373],
    "羊毛被/驼毛被":[15256,26252,34718,126048001],
    "蚕丝被":[15255,26241,34714,126050001],
    "被套":[15275,20987,34677,50001865,50008263],
    "羽绒/羽毛被":[15254,26172,34719,126046001],
    "休闲毯/毛毯/绒毯":[15273,17393,21160,26980,34720,34744,34745,50001871],
    "棉花被":[15258,26240,34715,126054001],
    "化纤被":[26239,34717,126052001,15257],
    "枕头/枕芯":[15259,15260,15261,15262,15263,15264,21104,26095,26161,26236,26237,27306,27324,34704,34705,34706,34707,34708,34709,34710,34711,34712,37712,37713,37714,37715,37716,37717,37718,37719,50002777,201304409,201309307,201309312,202146523,202149818],
    "床品套件/四件套/多件套":[15251,15252,15253,21273,34684,34685,34686,34687,290903,50008779,201305913,201310106,201663213],
    "床垫":[15274,21041,34695,34696,34697,34698,34699,34700,34701,34702,50008565,50019373]
   }

    return cate_dict

def get_brand_report(db,tbl,cate_dict,row_s,col_s,work_book,model_sheet_name,sheet_name):
    for i,cate in enumerate(cate_dict):
        where = """ WHERE  pkey >= '2023-08-01' AND pkey<'2024-08-01' and alias_all_bid not in (0,26) and cid in {cid_list}""".format(cid_list=cate_dict[cate])

        sql = get_brand_info(where=where, tbl=tbl)
        data = get_data(db, sql, as_dict=False)
        rr = pd.DataFrame(data, columns=["品牌","销量排名","销售额排名"])
        # if i == 0:
        #     dd = rr
        #     final_cols = cols = ["alias_all_bid","品牌","销量","销额","销量排名","销售额排名"]
        # else:
        #     dd = pd.merge(rr,dd,on=["销售额排名"], how='left', suffixes=('_total', '_item'))
        #     final_cols = cols + [''] + final_cols
        # data = np.array(dd).tolist()
    # print(data)
    # result = pd.DataFrame(rr, columns=final_cols)
    # result.to_excel(r'..\report\92447\\' + 'test.xlsx')
        work_book,col_s = write_xlsx(rr, row_s, col_s, work_book, model_sheet_name, sheet_name)
    return work_book

def get_cate10_report(db,tbl,cate_dict,row_s,col_s,work_book,model_sheet_name,sheet_name):
    for i,cate in enumerate(cate_dict):
        if cate_dict[cate] == [0]:
            where = """ WHERE  pkey >= '2023-08-01' AND pkey<'2024-08-01' and alias_all_bid in [22211,50647,68110,84613,24115,84516,989603,85792,143768,265393,51284,3626,50893,6786109,2833661,50860,5708533,2029107,720056,2101482,9488,6900990,10003,92794,1458202,143827,90748,84721,6963861,1933286,4495107,13630,51635,748250,6391699,2413988,2030094,1792664,760178,411118,6269251,2321041,85671,172,5904176,21409,220335,2638245,1190992,90785,69375,2413984,50729,4551931,51043,6208071,68552,147131,2529353,1345681,3224123,6314484,3242499,4379403,5905580,68128,3163182,653560,4820820,50726,26415,6522071,6510204,3638420,6581968,6894982,6764068,219747,250115,5175840,6729527,1039719,5392220,2695780,5075949,51512,6760922,5918699,3718928,657838,262877,2362554,540637,5787445,4411073,3227263,377854,6299865,84656,3852,3015947,1456927,6065371,68133,5194935,749508,5075962,6134788,2349967,3699024,6377258,91193,48476,91703,7009129,4886954,271885,1407699,5762449,747161,4254756,613,91858,143792,220309,3628811,68142,6040434,401474,2862,189679,7320688,6062982,3499163,2793749,35174,4421253,157620,877003,6054982,6090730,51694,51054,6304548,30649,8980,7034574,3587615,5281611,1265639,68462,4665225,90755,3675332,3671491,5279003,3935893,220620,6433403,1022073,6126122,7559176,332858,6286535,6565777,57796,7021834,3487758,3746449,3643832,7007875,6774148,3586605,3746594,3713969,3675338,6455456,893188,5912550,5181271,92190,219562,1993817,6008849,220358,6772675,7106604,1497497,50712,6076997,93535,90777,1792801,3770262,4748146,6421811,2827306,32563,1215550,5749514,219829,42464,5973222,143951,1907440,6654142,7194519,91760]"""
        else:
            where = """ WHERE  pkey >= '2023-08-01' AND pkey<'2024-08-01' and cid in {cid_list} and alias_all_bid in [22211,50647,68110,84613,24115,84516,989603,85792,143768,265393,51284,3626,50893,6786109,2833661,50860,5708533,2029107,720056,2101482,9488,6900990,10003,92794,1458202,143827,90748,84721,6963861,1933286,4495107,13630,51635,748250,6391699,2413988,2030094,1792664,760178,411118,6269251,2321041,85671,172,5904176,21409,220335,2638245,1190992,90785,69375,2413984,50729,4551931,51043,6208071,68552,147131,2529353,1345681,3224123,6314484,3242499,4379403,5905580,68128,3163182,653560,4820820,50726,26415,6522071,6510204,3638420,6581968,6894982,6764068,219747,250115,5175840,6729527,1039719,5392220,2695780,5075949,51512,6760922,5918699,3718928,657838,262877,2362554,540637,5787445,4411073,3227263,377854,6299865,84656,3852,3015947,1456927,6065371,68133,5194935,749508,5075962,6134788,2349967,3699024,6377258,91193,48476,91703,7009129,4886954,271885,1407699,5762449,747161,4254756,613,91858,143792,220309,3628811,68142,6040434,401474,2862,189679,7320688,6062982,3499163,2793749,35174,4421253,157620,877003,6054982,6090730,51694,51054,6304548,30649,8980,7034574,3587615,5281611,1265639,68462,4665225,90755,3675332,3671491,5279003,3935893,220620,6433403,1022073,6126122,7559176,332858,6286535,6565777,57796,7021834,3487758,3746449,3643832,7007875,6774148,3586605,3746594,3713969,3675338,6455456,893188,5912550,5181271,92190,219562,1993817,6008849,220358,6772675,7106604,1497497,50712,6076997,93535,90777,1792801,3770262,4748146,6421811,2827306,32563,1215550,5749514,219829,42464,5973222,143951,1907440,6654142,7194519,91760]""".format(cid_list=cate_dict[cate])

        sql = get_cate10_info(where=where, tbl=tbl)
        data = get_data(db, sql, as_dict=False)
        if i >=2:
            rr = pd.DataFrame(data, columns=["品牌","销量","销售额","rank"])[["销售额"]]
        else:
            rr = pd.DataFrame(data, columns=["品牌", "销量","销售额", "rank"])[["销量","销售额"]]
        # if i == 0:
        #     dd = rr
        #     final_cols = cols = ["alias_all_bid","品牌","销量","销额"]
        # else:
        #     dd = pd.merge(dd,rr,on=["alias_all_bid","品牌"], how='left', suffixes=('_total', '_item'))
        #     final_cols = cols +  final_cols
        # data = np.array(dd).tolist()
    # print(data)
    # result = pd.DataFrame(dd, columns=final_cols)
    # result.to_excel(r'..\report\92447\\' + 'test.xlsx')
        work_book,col_s = write_xlsx_nostyle(rr, row_s, col_s, work_book, model_sheet_name, sheet_name)
    return work_book

def style_map(rule):

    style_style = {
        'left_border':(7,1),
        'no_border':(7,2),
        'right_border':(7,3),
        'left_bottom_border':(59,1),
        'bottom_border': (59, 2),
        'right_bottom_border': (59, 3),
        'red_fill':(12,1)
    }
    if rule in style_style.keys():
        return style_style[rule]
    return

def get_style(wb,model_sheet_name,sheet_name,from_cell,to_cell):
    # 加载现有的 Excel 文件

    # 选择工作表
    ws = wb[model_sheet_name]
    # print(from_cell,to_cell)
    # 获取源单元格
    source_cell = ws.cell(from_cell[0],from_cell[1])

    # 创建新的样式对象
    new_font = Font(name=source_cell.font.name, size=source_cell.font.size,bold=source_cell.font.bold,italic=source_cell.font.italic,vertAlign=source_cell.font.vertAlign,underline=source_cell.font.underline,strike=source_cell.font.strike,color=source_cell.font.color)
    new_fill = PatternFill(fill_type=source_cell.fill.fill_type,start_color=source_cell.fill.start_color,end_color=source_cell.fill.end_color)
    new_border = Border(left=source_cell.border.left,right=source_cell.border.right, top=source_cell.border.top,bottom=source_cell.border.bottom,diagonal=source_cell.border.diagonal,diagonal_direction=source_cell.border.diagonal_direction,outline=source_cell.border.outline)
    new_alignment = Alignment(horizontal=source_cell.alignment.horizontal,vertical=source_cell.alignment.vertical,text_rotation=source_cell.alignment.text_rotation,wrap_text=source_cell.alignment.wrap_text,shrink_to_fit=source_cell.alignment.shrink_to_fit,indent=source_cell.alignment.indent)
    new_protection = Protection(locked=source_cell.protection.locked, hidden=source_cell.protection.hidden)


    # 应用新的样式对象到目标单元格
    ws = wb[sheet_name]
    target_cell = ws.cell(to_cell[0],to_cell[1])
    target_cell.font = new_font
    target_cell.fill = new_fill
    target_cell.border = new_border
    target_cell.alignment = new_alignment
    target_cell.number_format = source_cell.number_format
    target_cell.protection = new_protection
    return wb

def write_xlsx(data,row_s,col_s,work_book,model_sheet_name,sheet_name):

    max_row,max_col = data.shape
    data = np.array(data).tolist()
    for r in range(max_row):
        for c in range(max_col):
            if r == max_row - 1 and c==0:
                work_book = get_style(wb=work_book, model_sheet_name=model_sheet_name, sheet_name=sheet_name,from_cell=style_map('left_bottom_border'), to_cell=(row_s + r, col_s + c))
            elif r == max_row - 1 and c == 1:
                work_book = get_style(wb=work_book, model_sheet_name=model_sheet_name, sheet_name=sheet_name,from_cell=style_map('bottom_border'), to_cell=(row_s + r, col_s + c))
            elif r == max_row - 1 and c == 2:
                work_book = get_style(wb=work_book, model_sheet_name=model_sheet_name, sheet_name=sheet_name,from_cell=style_map('right_bottom_border'), to_cell=(row_s + r, col_s + c))
            elif str(data[r][c]) == '恒源祥':
                work_book = get_style(wb=work_book, model_sheet_name=model_sheet_name, sheet_name=sheet_name,from_cell=style_map('red_fill'), to_cell=(row_s + r, col_s + c))
            elif c == 0:
                work_book = get_style(wb=work_book,model_sheet_name=model_sheet_name,sheet_name=sheet_name,from_cell=style_map('left_border'),to_cell=(row_s+r,col_s+c))
            elif c == 1:
                work_book = get_style(wb=work_book,model_sheet_name=model_sheet_name,sheet_name=sheet_name, from_cell=style_map('no_border'),to_cell=(row_s + r, col_s + c))
            elif c == 2:
                work_book = get_style(wb=work_book,model_sheet_name=model_sheet_name,sheet_name=sheet_name, from_cell=style_map('right_border'),to_cell=(row_s + r, col_s + c))
            else:
                pass
            sheet = work_book[sheet_name]
            sheet.cell(row_s+r,col_s+c).value = data[r][c]
    return work_book,max_col+col_s+1

def write_xlsx_nostyle(data,row_s,col_s,work_book,model_sheet_name,sheet_name):

    max_row,max_col = data.shape
    data = np.array(data).tolist()
    for r in range(max_row):
        for c in range(max_col):
            if col_s>=7:
                sheet = work_book[sheet_name]
                sheet.cell(row_s+r,col_s+c+11).value = data[r][c]/1000000
                try:
                    sheet.cell(row_s + r, col_s + c).value = data[r][c]/sheet.cell(row_s+r,4).value
                except:
                    sheet.cell(row_s + r, col_s + c).value = '计算异常'

                try:
                    sheet.cell(row_s + r, 17).value = sheet.cell(row_s+r,6).value/sheet.cell(row_s+r,4).value
                except:
                    sheet.cell(row_s + r, col_s + c).value = '计算异常'
            else:
                sheet = work_book[sheet_name]
                sheet.cell(row_s + r, col_s + c).value = data[r][c]
    return work_book,max_col+col_s

def main():
    #######################################  取报告可变的全局参数，可以任意改变  #############################################

    tbl = 'sop.entity_prod_92447_A'  #从哪一张e表取报告
    template = '恒源祥-认证数据（原始）-240809V2-更新.xlsx' #报告魔板工作簿
    output = '恒源祥-认证数据-20240917-更新.xlsx' #报告最终输出的工作簿名称
    model_sheet_name = '恒源祥重点类目排名情况（模版）' #报告模板在哪一个sheet
    sheet_name = '恒源祥重点类目排名情况' #报告要保存在哪一个sheet
    # group_date = 'toStartOfMonth(pkey) AS Gmonth' # 此为分月取报告时用，默认分年，也可替换其它时间维度如：分季、分fy、mat等，只需要改变AS前面即可
    # group_date = 'year(pkey) AS Gmonth'  # 此为分年取报告时用，默认分年，也可替换其它时间维度如：分季、分fy、mat等，只需要改变AS前面即可
    db = connect_clickhouse('chsop')
    work_book = load_workbook(r'..\report\92447\\' + template)
    ####################################################################################################################

    ##############限定sku的范围，固定的sku可用列表直接限定，现在默认通过get_sku_list去取排除范围外的18个sku#########################

    # sku_list = ['FACE POWDER','LIPOSOME  ADVANCED REPAIR SERUM','PRIME LATTE','SUN SHELTER MULTI PROTECTION','ULTIMUNE']
    cate_dict = get_cate_dict()
    cate_dict2 = get_cate_dict2()
    ####################################################################################################################

    #######################################  获取sku报告的核心代码  #######################################################
    work_book = get_brand_report(db=db, tbl=tbl, cate_dict=cate_dict,row_s=7,col_s=1,work_book=work_book,model_sheet_name=model_sheet_name,sheet_name=sheet_name)
    get_cate10_report(db=db, tbl=tbl, cate_dict=cate_dict2,row_s=3,col_s=3,work_book=work_book,model_sheet_name='9个重点类目类目结构占比',sheet_name='9个重点类目类目结构占比')
    ################################################# 保存sku报告  ######################################################
    work_book.save(r'..\report\92447\\' + output)
    #######################################  获取sku报告的核心代码  #######################################################

if __name__ == '__main__':
    main()
