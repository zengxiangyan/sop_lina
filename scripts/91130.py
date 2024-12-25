# -*- coding: utf-8 -*-
from common import source_cn,source_en,source_link,cid_name
from db import connect_clickhouse
from os.path import abspath, join, dirname
import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Fill, Border, Alignment, Protection
sys.path.insert(0, join(abspath(dirname(__file__)), '../../'))
report_path = r'{path}'.format(path=sys.path[0])

def get_sku_info(where,tbl,group_date=False):
    if group_date == False:
        group_date = 'year(pkey) AS Gmonth'

    sql="""
        select 'DECORTE' "対象メーカー"
        ,"対象製品"          
        ,"spSKU容量" "容量"
        ,if("合并行" = 7,'合计',"平台") "ECプラットフォーム"
        ,Gmonth "年月"
        ,if("合并行" = 7,'合计',"sp店铺分类") "店舗タイプ"
        ,if("合并行" = 7,'合计',STORE_NAME) "店舗名"
        ,sales_value "金額規模（元）"
        ,sales_volume "販売量規模（個数）"
        ,Price_perUnit "取引個数単価（元/1本あたり）"
        --,image
        ,sid        
        --,if("合并行" = 7,'NA',Price_perUnit) "取引個数単価（元/1本あたり）"
        --, "最低取引個数単価（元/1本あたり）"
        from(
                select 
                "spSKU名" ||"spSKU容量" "対象製品" 
                ,"spSKU容量"
                ,case when source = 1 and shop_type < 20 and shop_type != 9 then 'Taobao'
                         when source = 1 and shop_type > 20 or source = 1 and shop_type = 9 then 'Tmall'
                     when source = 2 then 'JD'
                     when source = 5 then 'kaola'        
                     when source = 6 then 'suning' 
                     when source = 8 then 'Pin duo duo'
                     when source = 9 then 'jiuxian'
                     when source = 11 then 'Douyin'             
                     else '其他' end as "平台"
                --, toStartOfMonth(pkey) AS Gmonth
                --, year(pkey) AS Gmonth
                ,{group_date}
                ,if(sid in (107428076,192151573,1000223736,1000427454),'旗艦店','非旗艦店') "sp店铺分类"
                ,dictGet('all_shop', 'title', tuple(toUInt8(source), toUInt32(sid))) AS STORE_NAME
                ,grouping("平台","sp店铺分类",STORE_NAME) "合并行" --揉在一起不显示的行，2^0+2^1+2^2=7
                ,sum(sales)/100 AS sales_value
                ,sum((num)*toFloat64OrZero("spSKU件数")) AS sales_volume
                ,sales_value/sales_volume as Price_perUnit
            ,argMin(sales/num/100,date) as "最低取引個数単価（元/1本あたり）"
            ,argMax(img,date)as image
            ,sid
                FROM {tbl}
                {where}
                group by grouping sets(("対象製品" ,"spSKU容量","平台",Gmonth,"sp店铺分类",STORE_NAME,sid),("対象製品" ,"spSKU容量",Gmonth)) --做汇总行，前面的括号是其他的所有分组值，后面括号里的三列是要显示的数据。
                )
        order by "対象製品","spSKU容量", "ECプラットフォーム" = '合计',sales_value desc --平台=合计放最后一行
        """.format(tbl=tbl,where=where,group_date=group_date)

    return sql

def get_item_info(db,where,tbl,group_date=False):
    if group_date == False:
        group_date = 'year(pkey) AS Gmonth'

    sql="""
            select 'DECORTE' "対象メーカー"
            ,"対象製品"          
            ,"spSKU容量" "容量"
            ,"平台" "ECプラットフォーム"
            ,Gmonth "年月"
            ,"sp店铺分类" "店舗タイプ"
            ,STORE_NAME "店舗名"
            ,item_sales "金額規模（元）"
            ,item_volume "販売量規模（個数）"
            ,item_Price_perUnit "取引個数単価（元/1本あたり）"
            ,sid        
            from(
                select 
                "spSKU名" ||"spSKU容量" "対象製品" 
                ,"spSKU容量"
                ,case when source = 1 and shop_type < 20 and shop_type != 9 then 'Taobao'
                         when source = 1 and shop_type > 20 or source = 1 and shop_type = 9 then 'Tmall'
                     when source = 2 then 'JD'
                     when source = 5 then 'kaola'        
                     when source = 6 then 'suning' 
                     when source = 8 then 'Pin duo duo'
                     when source = 9 then 'jiuxian'
                     when source = 11 then 'Douyin'             
                     else '其他' end as "平台"
                --, toStartOfMonth(pkey) AS Gmonth
                --, year(pkey) AS Gmonth
                ,{group_date}
                ,if(sid in (107428076,192151573,1000223736,1000427454),'旗艦店','非旗艦店') "sp店铺分类"
                ,dictGet('all_shop', 'title', tuple(toUInt8(source), toUInt32(sid))) AS STORE_NAME
                ,sum(sales)/100 AS item_sales
                ,sum((num)*toFloat64OrZero("spSKU件数")) AS item_volume
                ,item_sales/item_volume as item_Price_perUnit
            ,item_id
            ,sid
                FROM {tbl}
                {where}
                group by "対象製品" ,"spSKU容量","平台",Gmonth,"sp店铺分类",STORE_NAME,sid,item_id) --做汇总行，前面的括号是其他的所有分组值，后面括号里的三列是要显示的数据。
            ORDER BY "取引個数単価（元/1本あたり）"
            LIMIT 1 BY "対象製品" ,"spSKU容量","平台",Gmonth,"sp店铺分类",STORE_NAME,sid
        """.format(tbl=tbl,where=where,group_date=group_date)
    item_price = get_data(db,sql)
    item_price = pd.DataFrame(item_price)
    return item_price

def get_data(db,sql,as_dict=True):

    data = db.query_all(sql,as_dict=as_dict)

    return data

def get_sku_list(db,tbl):
    sql = """
            select 
            "spSKU名" ||"spSKU容量" "対象製品" ,SUM(sales)`销售额`
            FROM {tbl}
            WHERE  pkey>='2023-01-01' and pkey<'2024-01-01'
                and alias_all_bid=218724 --限制黛珂
                and "source" in (1,2,11) --4个平台
                and "sp是否人工答题" in ('前后月覆盖','出题宝贝')              
                and "spSKU名" not in ('LIPOSOME  ADVANCED REPAIR SERUM','____其它','___待客户确认(同规格限定/替换)')
            GROUP BY "対象製品"
            ORDER BY `销售额` DESC;
            """.format(tbl=tbl)
    sku_list = db.query_all(sql,as_dict=False)

    return [sku[0] for sku in sku_list]

def get_sku_report(db,tbl,sku_list,row_s,col_s,work_book,model_sheet_name,sheet_name,group_date=False):
    for i,sku in enumerate(sku_list):
        where = """ WHERE  pkey>='2023-01-01' and pkey<'2024-01-01'
                and alias_all_bid=218724 --限制黛珂
                and "平台" in ('JD','Douyin','Tmall','Taobao') --4个平台
                and "sp是否人工答题" in ('前后月覆盖','出题宝贝')    
                and ("spSKU名" ||"spSKU容量") = '{sku}' """.format(sku=sku)
        price_data = get_item_info(db=db,where=where,group_date=group_date,tbl=tbl)
        sql = get_sku_info(where=where,group_date=group_date, tbl=tbl)
        data = get_data(db, sql, as_dict=False)
        rr = pd.DataFrame(data, columns=['対象メーカー','対象製品','容量','ECプラットフォーム','年月','店舗タイプ', '店舗名', '金額規模（元）','販売量規模（個数）', '取引個数単価（元/1本あたり）', 'sid'])
        rr = pd.merge(rr, price_data, on=['対象メーカー','対象製品','容量','ECプラットフォーム','年月','店舗タイプ', '店舗名','sid'], how='left', suffixes=('_total', '_item'))
        data = np.array(rr).tolist()
        rr = [[sku,'','','','','','','','','','','','',''],['202307','','','','','','','','','','','','',''],['対象メーカー','対象製品','容量','ECプラットフォーム','年月','店舗タイプ', '店舗名', '金額規模（元）','販売量規模（個数）', '取引個数単価（元/1本あたり）', 'sid','item_sales','item_volume','item_Price_perUnit']] + data + [['', '', '', '', '', '','','','','',''], ['', '', '', '', '', '','','','','','']]
        rr = pd.DataFrame(rr, columns=['対象メーカー','対象製品','容量','ECプラットフォーム','年月','店舗タイプ', '店舗名', '金額規模（元）','販売量規模（個数）', '取引個数単価（元/1本あたり）', 'sid','item_sales','item_volume','item_Price_perUnit'])
        work_book,row_s = write_xlsx(rr, row_s, col_s, work_book, model_sheet_name, sheet_name)
    return work_book

def style_map(rule):

    style_style = {
        'sku_title':(3,1),
        'date_title':(4,1),
        'col_title':(5,1),
        'total_row':(18,1),
        'normal':(6,1)
    }
    if rule in style_style.keys():
        return style_style[rule]
    return

def get_style(wb,model_sheet_name,sheet_name,from_cell,to_cell):
    # 加载现有的 Excel 文件

    # 选择工作表
    ws = wb[model_sheet_name]

    # 获取源单元格
    source_cell = ws.cell(from_cell[0],from_cell[1])

    # 创建新的样式对象
    new_font = Font(name=source_cell.font.name, size=source_cell.font.size,bold=source_cell.font.bold,italic=source_cell.font.italic,vertAlign=source_cell.font.vertAlign,underline=source_cell.font.underline,strike=source_cell.font.strike,color=source_cell.font.color)
    new_fill = PatternFill(fill_type=source_cell.fill.fill_type,start_color=source_cell.fill.start_color,end_color=source_cell.fill.end_color)
    new_border = Border(left=source_cell.border.left,right=source_cell.border.right, top=source_cell.border.top,bottom=source_cell.border.bottom,diagonal=source_cell.border.diagonal,diagonal_direction=source_cell.border.diagonal_direction,outline=source_cell.border.outline)
    new_alignment = Alignment(horizontal=source_cell.alignment.horizontal,vertical=source_cell.alignment.vertical,text_rotation=source_cell.alignment.text_rotation,wrap_text=source_cell.alignment.wrap_text,shrink_to_fit=source_cell.alignment.shrink_to_fit,indent=source_cell.alignment.indent)
    new_protection = Protection(locked=source_cell.protection.locked, hidden=source_cell.protection.hidden)

    # 应用新的样式对象到目标单元格
    ws = wb[sheet_name]
    target_cell = ws.cell(to_cell[0],to_cell[1])
    target_cell.font = new_font
    target_cell.fill = new_fill
    target_cell.border = new_border
    target_cell.alignment = new_alignment
    target_cell.number_format = source_cell.number_format
    target_cell.protection = new_protection
    return wb

def write_xlsx(data,row_s,col_s,work_book,model_sheet_name,sheet_name):

    max_row,max_col = data.shape
    data = np.array(data).tolist()
    for r in range(max_row):
        for c in range(max_col):
            if r == 0:
                work_book = get_style(wb=work_book,model_sheet_name=model_sheet_name,sheet_name=sheet_name,from_cell=style_map('sku_title'),to_cell=(row_s+r,col_s+c))
            elif r == 1:
                work_book = get_style(wb=work_book,model_sheet_name=model_sheet_name,sheet_name=sheet_name, from_cell=style_map('date_title'),to_cell=(row_s + r, col_s + c))
            elif r == 2:
                work_book = get_style(wb=work_book,model_sheet_name=model_sheet_name,sheet_name=sheet_name, from_cell=style_map('col_title'),to_cell=(row_s + r, col_s + c))
            elif r == max_row-3:
                work_book = get_style(wb=work_book,model_sheet_name=model_sheet_name,sheet_name=sheet_name, from_cell=style_map('total_row'),to_cell=(row_s + r, col_s + c))
            elif r < max_row-3:
                work_book = get_style(wb=work_book,model_sheet_name=model_sheet_name,sheet_name=sheet_name, from_cell=style_map('normal'),to_cell=(row_s + r, col_s + c))
            else:
                pass
            sheet = work_book[sheet_name]
            sheet.cell(row_s+r,col_s+c).value = data[r][c]
    return work_book,max_row+row_s

def main():
    #######################################  取报告可变的全局参数，可以任意改变  #############################################

    tbl = 'sop_e.entity_prod_92312_E_2024'  #从哪一张e表取报告
    template = 'コーセー様_納品データ_231120.xlsx' #报告魔板工作簿
    output = 'コーセー様_納品データ_231120 output.xlsx' #报告最终输出的工作簿名称
    model_sheet_name = '模板' #报告模板在哪一个sheet
    sheet_name = '报告' #报告要保存在哪一个sheet
    # group_date = 'toStartOfMonth(pkey) AS Gmonth' # 此为分月取报告时用，默认分年，也可替换其它时间维度如：分季、分fy、mat等，只需要改变AS前面即可
    group_date = 'year(pkey) AS Gmonth'  # 此为分年取报告时用，默认分年，也可替换其它时间维度如：分季、分fy、mat等，只需要改变AS前面即可
    db = connect_clickhouse('chsop')
    work_book = load_workbook(r'..\report\92312\\' + template)
    ####################################################################################################################

    ##############限定sku的范围，固定的sku可用列表直接限定，现在默认通过get_sku_list去取排除范围外的18个sku#########################

    # sku_list = ['FACE POWDER','LIPOSOME  ADVANCED REPAIR SERUM','PRIME LATTE','SUN SHELTER MULTI PROTECTION','ULTIMUNE']
    sku_list = get_sku_list(db,tbl)

    ####################################################################################################################

    #######################################  获取sku报告的核心代码  #######################################################
    work_book = get_sku_report(db=db, tbl=tbl, sku_list=sku_list,row_s=3,col_s=1,work_book=work_book,model_sheet_name=model_sheet_name,sheet_name=sheet_name,group_date=group_date)

    ################################################# 保存sku报告  ######################################################
    # work_book.save(r'..\report\92312\\' + output)
    #######################################  获取sku报告的核心代码  #######################################################

if __name__ == '__main__':
    main()