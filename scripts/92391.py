# -*- coding: utf-8 -*-
from common import source_cn,source_en,source_link,cid_name
from db import connect_clickhouse
from os.path import abspath, join, dirname
import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Fill, Border, Alignment, Protection
sys.path.insert(0, join(abspath(dirname(__file__)), '../../'))
report_path = r'{path}'.format(path=sys.path[0])

def get_rowdata_info(where,tbl,limit=10):

    sql="""
            SELECT 
                case when source = 11 then 'douyin' else '其他' end as platform,
                argMax(uuid2,num)"no",toStartOfMonth(pkey)"time",argMax(name,num)"final_name",
                case when source = 11 then CONCAT('https://haohuo.jinritemai.com/views/product/detail?id=',item_id) else '其他' end as url,
                ''"shopnameold",''"shopname",''"ShopType2",'' "FS ShopType",SUM(num)"unit",SUM(sales)/(SUM(num)*100) "price",SUM(sales)/100 AS total_sales,cid,
                ''"User",''"Manufacturer",''"Division",''"Selectivity",''"BrandLRL",''"Overseas Shop Type",
                sid,dictGet('all_shop', 'title', tuple(toUInt8(source), toUInt32(sid))) AS storename,dictGetOrDefault('all_brand', 'name', tuple(toUInt32(alias_all_bid)), '') as `品牌名`
            from {tbl}
            {where}
            GROUP BY platform,"time",url,cid,sid,storename,alias_all_bid
            ORDER BY total_sales desc;
        """.format(tbl=tbl,where=where)

    return sql

def get_data(db,sql,as_dict=True):

    data = db.query_all(sql,as_dict=as_dict)

    return data

def get_rowdata_report(db,tbl,where):
    sql = get_rowdata_info(where=where, tbl=tbl)
    data = get_data(db, sql, as_dict=True)
    rr = pd.DataFrame(data)
    return rr

def write_xlsx(data,row_s,col_s,work_book,sheet_name):

    sheet = work_book[sheet_name]
    # print(sheet.cell(1,1).style)
    max_row,max_col = data.shape
    data = np.array(data).tolist()
    for r in range(max_row):
        for c in range(max_col):
            sheet.cell(row_s+r,col_s+c).value = data[r][c]
    return work_book

def main():
    tbl = 'sop.entity_prod_92391_A'
    where = """
            WHERE `date` >='2022-01-01'
            and `date` <'2024-01-01'
            and num!=0
            and (match(name, '面膜|面贴膜') or match(toString(`trade_props.value`) , '面膜|面贴膜') or cid in (33361,33363,33364))
            """
    json = {
        'start_date': '2022-01-01',
        'end_date': '2024-01-01',
        'eid': '92391',
        'table': 'entity_prod_92391_A',
        'cid': [],
    }
    template = '2311月导入版_20240105.xlsx'
    output = '抖音医疗健康2022-2023全年分月_20240311.xlsx'
    sheet_name = '导入版'
    db = connect_clickhouse('chsop')
    work_book = load_workbook(r'..\report\92391\\' + template)
    # 获取sku报告
    rr = get_rowdata_report(db=db, tbl=tbl,where=where)
    print(rr)
    work_book = write_xlsx(rr, 2, 1, work_book, sheet_name)
    # # 保存工作簿
    work_book.save(r'..\report\92391\\' + output)


if __name__ == '__main__':
    main()