from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection

# 加载现有的 Excel 文件
wb = load_workbook('example.xlsx')

# 选择工作表
ws = wb.active

# 获取源单元格
source_cell = ws['A18']

# 创建新的样式对象
new_font = Font(name=source_cell.font.name,size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                vertAlign=source_cell.font.vertAlign,
                underline=source_cell.font.underline,
                strike=source_cell.font.strike,
                color=source_cell.font.color)

new_fill = PatternFill(fill_type=source_cell.fill.fill_type,
                       start_color=source_cell.fill.start_color,
                       end_color=source_cell.fill.end_color)

new_border = Border(left=source_cell.border.left,
                    right=source_cell.border.right,
                    top=source_cell.border.top,
                    bottom=source_cell.border.bottom,
                    diagonal=source_cell.border.diagonal,
                    diagonal_direction=source_cell.border.diagonal_direction,
                    outline=source_cell.border.outline)

new_alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                          vertical=source_cell.alignment.vertical,
                          text_rotation=source_cell.alignment.text_rotation,
                          wrap_text=source_cell.alignment.wrap_text,
                          shrink_to_fit=source_cell.alignment.shrink_to_fit,
                          indent=source_cell.alignment.indent)

new_protection = Protection(locked=source_cell.protection.locked,
                            hidden=source_cell.protection.hidden)

# 应用新的样式对象到目标单元格
target_cell = ws.cell(5,1)
target_cell.font = new_font
target_cell.fill = new_fill
target_cell.border = new_border
target_cell.alignment = new_alignment
target_cell.number_format = source_cell.number_format
target_cell.protection = new_protection

# 保存工作簿
wb.save('example.xlsx')

# 不要忘记关闭工作簿
wb.close()
