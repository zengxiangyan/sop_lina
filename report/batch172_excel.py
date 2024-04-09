# -*- coding: utf-8 -*-
import sys
from os.path import abspath, join, dirname
from datetime import datetime

import pandas as pd
from dateutil.relativedelta import relativedelta
import asyncio
from openpyxl import load_workbook
sys.path.insert(0, join(abspath(dirname(__file__)), '../'))

from sop.connect_clickhouse import async_connect

def sql_date_info(d0,d1,d2,d3,d4):
    #2023-07-04新模板
    sql_list = [f"""SELECT toStartOfMonth(pkey) AS Gmonth,`sp厂商`,`sp子品类`,`sp产品品牌`,`sp子品牌`,`sp件数`,item_id,trade_props.name,trade_props.value as `交易属性`,
    (model*sum(if(model!=0,num,0)))/1000000 as volume,
    toFloat64OrZero(`sp总规格`) as model,
    case when Gmonth >= '2021-01-01' AND Gmonth < '2022-01-01' then 'FY21'
         when Gmonth >= '2022-01-01' AND Gmonth < '2023-01-01' then 'FY22'
         when Gmonth >= '2023-01-01' AND Gmonth < '2024-01-01' then 'FY23'           
    else '其他' end as FY,
    case when Gmonth >= '{d1}' AND Gmonth < '{d2}' then 'YTD22'
         when Gmonth >= '{d3}' AND Gmonth < '{d4}' then 'YTD23'           
    else '其他' end as YTD,
    case when `sp子品类`in ['婴儿牛奶粉','婴儿特殊配方粉','儿童牛奶粉'] THEN '限制' ELSE '否' END AS `是否限制子品类`,
    case when `sp厂商` = 'Frieslandcampina/荷兰皇家菲仕兰 ' AND `sp产品品牌` = 'Friso Prestige/皇家美素佳儿' then 'up'
         when `sp厂商` = 'Danone/达能 ' AND `sp产品品牌` = 'Aptamil/爱他美' AND `sp子品牌` = '卓萃 ' then 'up'
         when `sp厂商` = 'Firmus/飞鹤 ' AND `sp产品品牌` = '星飞帆' AND `sp子品牌` IN ['','卓睿'] then 'up'
         when `sp厂商` = 'Firmus/飞鹤  ' AND `sp产品品牌` = '臻稚' AND `sp子品牌` = '有机 ' then 'up'
         when `sp厂商` = 'Yili/伊利 ' AND `sp产品品牌` = '金领冠Pro-Kido' AND `sp子品牌` = '珍护 ' then 'up'
         when `sp厂商` = 'A2/The A2 Milk Company' AND `sp产品品牌` = 'A2' AND `sp子品牌` in ['a2至初','A2至初'] then 'up'
         when `sp厂商` = 'H&Hgroup/健合集团' AND `sp产品品牌` = 'Biostime/合生元' AND `sp子品牌` = '派星 ' then 'up'
         when `sp厂商` = 'Wyeth/惠氏' AND `sp产品品牌` = 'Illuma/启赋' AND `sp子品牌` IN ['蓝钻 ','蕴淳','Organic','铂金'] then 'up'
         when `sp厂商` = 'Mead Johnson/美赞臣 ' AND `sp产品品牌` = 'Enfinitas/蓝臻 ' then 'up'
    else '其他' end as `是否up`,
    `sp适用人群（段数）`,SUM(num) AS total_num, SUM(sales)/100000000 AS total_sales  
    FROM sop_e.entity_prod_91357_E_meisu WHERE (date >= '{d0}' AND date < '{d4}'
    AND source*100+shop_type IN [123,126,121,125,211,212] 
    AND cid in (211104,201284105,7052,31762) 
    AND `sp是否无效链接`!='无效链接') 
    GROUP BY Gmonth,item_id,`sp件数`,model,`sp子品类`,`sp适用人群（段数）`,`sp厂商`,`sp产品品牌`,`sp子品牌`,`交易属性`,trade_props.name"""]
    return sql_list

def clean(mydata):
    # 原定up规则
    mydata.loc[(mydata['sp厂商'] == 'Firmus/飞鹤') & (mydata['sp子品牌'] == '卓睿') & (mydata['sp产品品牌'] == '星飞帆'), '是否up'] = 'up'
    mydata.loc[(mydata['sp厂商'] == 'Firmus/飞鹤') & (mydata['sp子品牌'] == '') & (mydata['sp产品品牌'] == '星飞帆'), '是否up'] = 'up'
    mydata.loc[(mydata['sp厂商'] == 'Frieslandcampina/荷兰皇家菲仕兰') & (mydata['sp产品品牌'] == 'Friso Prestige/皇家美素佳儿'), '是否up'] = 'up'
    mydata.loc[(mydata['sp厂商'] == 'Danone/达能') & (mydata['sp子品牌'] == '卓萃') & (mydata['sp产品品牌'] == 'Aptamil/爱他美'), '是否up'] = 'up'
    mydata.loc[(mydata['sp厂商'] == 'Firmus/飞鹤') & (mydata['sp子品牌'] == '有机') & (mydata['sp产品品牌'] == '臻稚'), '是否up'] = 'up'
    mydata.loc[(mydata['sp厂商'] == 'Yili/伊利') & (mydata['sp子品牌'] == '珍护') & (mydata['sp产品品牌'] == '金领冠Pro-Kido'), '是否up'] = 'up'
    mydata.loc[(mydata['sp厂商'] == 'A2/The A2 Milk Company') & (mydata['sp子品牌'] == 'a2至初') & (mydata['sp产品品牌'] == 'A2'), '是否up'] = 'up'
    mydata.loc[(mydata['sp厂商'] == 'A2/The A2 Milk Company') & (mydata['sp子品牌'] == 'A2至初') & (mydata['sp产品品牌'] == 'A2'), '是否up'] = 'up'
    mydata.loc[(mydata['sp厂商'] == 'H&Hgroup/健合集团') & (mydata['sp子品牌'] == '派星') & (mydata['sp产品品牌'] == 'Biostime/合生元'), '是否up'] = 'up'
    mydata.loc[(mydata['sp厂商'] == 'Wyeth/惠氏') & (mydata['sp子品牌'].isin(['蓝钻', '蕴淳', 'Organic', '铂金'])) & (mydata['sp产品品牌'] == 'Illuma/启赋'), '是否up'] = 'up'
    mydata.loc[(mydata['sp厂商'] == 'Mead Johnson/美赞臣') & (mydata['sp产品品牌'] == 'Enfinitas/蓝臻'), '是否up'] = 'up'

    # 2023-07-04新增SP规则
    mydata['是否sp'] = ''
    mydata.loc[(mydata['sp厂商'] == 'Mead Johnson/美赞臣') & (mydata['sp产品品牌'].isin(['Enfamil A+/安婴儿 A+', 'Enfamil A+', 'Enfamil A2', 'Enfamil AⅡ'])), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Mead Johnson/美赞臣') & (mydata['sp产品品牌'] == 'Enfagrow A+/安儿宝 A+'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Mead Johnson/美赞臣') & (mydata['sp产品品牌'] == 'Enfa A+'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Mead Johnson/美赞臣') & (mydata['sp子品牌'] == '卓智') & (mydata['sp产品品牌'] == 'Nutri Power/学优力'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Wyeth/惠氏') & (mydata['sp子品牌'].isin(['Ultima/铂臻', 'Ultima Pdf/铂臻蔼而嘉'])) & (mydata['sp产品品牌'] == 'S-26'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Frieslandcampina/荷兰皇家菲仕兰') & (mydata['sp子品牌'] == '源悦') & (mydata['sp产品品牌'] == 'Friso/美素佳儿'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Firmus/飞鹤') & (mydata['sp产品品牌'] == '超级飞帆'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Firmus/飞鹤') & (mydata['sp产品品牌'] == '舒贝诺'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Firmus/飞鹤') & (mydata['sp产品品牌'] == '星阶优护'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Firmus/飞鹤') & (mydata['sp产品品牌'].isin(['臻爱飞帆', '精粹', '茁然', '飞睿'])), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'H&Hgroup/健合集团') & (mydata['sp子品牌'] == '贝塔星') & (mydata['sp产品品牌'] == 'Biostime/合生元'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'H&Hgroup/健合集团') & (mydata['sp子品牌'].isin(['满乐', '贝素贝加', '素加', 'Terroir/沃蓝', '儿童成长'])) & (mydata['sp产品品牌'] == 'Biostime/合生元'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'].isin(['帜亲', ''])) & (mydata['sp产品品牌'] == 'Banner Dairy/旗帜'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'] == '益佳') & (mydata['sp产品品牌'] == 'Banner Dairy/旗帜'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'] == '诠维爱') & (mydata['sp产品品牌'] == '小小鲁班'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'] == '诠维爱aii') & (mydata['sp产品品牌'] == '小小鲁班'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'] == '恬适') & (mydata['sp产品品牌'] == '君乐宝'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'].isin(['乐臻', '乐星', '淳护', '澳力高'])) & (mydata['sp产品品牌'] == '君乐宝'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Yili/伊利') & (mydata['sp子品牌'] == '菁护') & (mydata['sp产品品牌'] == '金领冠Pro-Kido'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Yili/伊利') & (mydata['sp子品牌'] == '睿护') & (mydata['sp产品品牌'] == '金领冠Pro-Kido'), '是否sp'] = 'sp'
    mydata.loc[(mydata['sp厂商'] == 'Yili/伊利') & (mydata['sp子品牌'] == '有机') & (mydata['sp产品品牌'] == '金领冠Pro-Kido'), '是否sp'] = 'sp'

    # 2023-07-04新增P规则
    mydata['是否p'] = ''
    mydata.loc[(mydata['sp厂商'] == 'Mead Johnson/美赞臣') & (mydata['sp产品品牌'] == 'Enfamil A+/铂睿 A+'), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'Mead Johnson/美赞臣') & (mydata['sp产品品牌'].isin(['Enfakid A+/安儿健 A+', 'Enfaschool A+/安学健 A+'])), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'Mead Johnson/美赞臣') & (mydata['sp子品牌'].isin(['', '营护'])) & (mydata['sp产品品牌'] == 'Nutri Power/学优力'), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'Wyeth/惠氏') & (mydata['sp子品牌'] == 'Gold') & (mydata['sp产品品牌'] == 'S-26'), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'Wyeth/惠氏') & (mydata['sp子品牌'].isin(['膳儿加', '学儿乐', '幼儿乐'])) & (mydata['sp产品品牌'] == 'S-26'), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'Frieslandcampina/荷兰皇家菲仕兰') & (mydata['sp子品牌'].isin(['金装', ''])) & (mydata['sp产品品牌'] == 'Friso/美素佳儿'), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'Firmus/飞鹤') & (mydata['sp产品品牌'] == '飞帆'), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'H&Hgroup/健合集团') & (mydata['sp子品牌'] == '阿尔法星') & (mydata['sp产品品牌'] == 'Biostime/合生元'), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'].isin(['诠护爱', '诠力爱', ''])) & (mydata['sp产品品牌'] == '小小鲁班'), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'].isin(['乐铂', '乐铂k2'])) & (mydata['sp产品品牌'] == '君乐宝'), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'] == '至臻') & (mydata['sp产品品牌'] == '君乐宝'), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'] == '小旗才') & (mydata['sp产品品牌'] == 'Banner Dairy/旗帜'), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'].isin(['乐纯', '乐畅'])) & (mydata['sp产品品牌'] == '君乐宝'), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'Yili/伊利') & (mydata['sp子品牌'].isin(['塞纳牧', '赋能', '呵护', '悠滋小羊', '育护', '珍护菁蕴', ''])) & (mydata['sp产品品牌'] == '金领冠Pro-Kido'), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'Yili/伊利') & (mydata['sp子品牌'].isin(['倍冠', '赋能星'])) & (mydata['sp产品品牌'] == '伊利'), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'Yili/伊利') & (mydata['sp产品品牌'] == 'QQ星'), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'Danone/达能') & (mydata['sp子品牌'] != '卓萃') & (mydata['sp产品品牌'] == 'Aptamil/爱他美'), '是否p'] = 'p'
    mydata.loc[(mydata['sp厂商'] == 'Danone/达能') & (mydata['sp产品品牌'].isin(['Nutrilon/诺优能', 'Nutrilon/诺贝能'])), '是否p'] = 'p'

    mydata['rule'] = ''
    mydata.loc[(mydata['sp厂商'] == 'Firmus/飞鹤') & (mydata['sp子品牌'] == '卓睿') & (mydata['sp产品品牌'] == '星飞帆'), 'rule'] = 'r3'
    mydata.loc[(mydata['sp厂商'] == 'Firmus/飞鹤') & (mydata['sp子品牌'] == '') & (mydata['sp产品品牌'] == '星飞帆'), 'rule'] = 'r3'
    mydata.loc[(mydata['sp厂商'] == 'Frieslandcampina/荷兰皇家菲仕兰') & (mydata['sp产品品牌'] == 'Friso Prestige/皇家美素佳儿'), 'rule'] = 'r1'
    mydata.loc[(mydata['sp厂商'] == 'Danone/达能') & (mydata['sp子品牌'] == '卓萃') & (mydata['sp产品品牌'] == 'Aptamil/爱他美'), 'rule'] = 'r2'
    mydata.loc[(mydata['sp厂商'] == 'Firmus/飞鹤') & (mydata['sp子品牌'] == '有机') & (mydata['sp产品品牌'] == '臻稚'), 'rule'] = 'r4'
    mydata.loc[(mydata['sp厂商'] == 'Yili/伊利') & (mydata['sp子品牌'] == '珍护') & (mydata['sp产品品牌'] == '金领冠Pro-Kido'), 'rule'] = 'r5'
    mydata.loc[(mydata['sp厂商'] == 'A2/The A2 Milk Company') & (mydata['sp子品牌'] == 'a2至初') & (mydata['sp产品品牌'] == 'A2'), 'rule'] = 'r6'
    mydata.loc[(mydata['sp厂商'] == 'A2/The A2 Milk Company') & (mydata['sp子品牌'] == 'A2至初') & (mydata['sp产品品牌'] == 'A2'), 'rule'] = 'r6'
    mydata.loc[(mydata['sp厂商'] == 'H&Hgroup/健合集团') & (mydata['sp子品牌'] == '派星') & (mydata['sp产品品牌'] == 'Biostime/合生元'), 'rule'] = 'r7'
    mydata.loc[(mydata['sp厂商'] == 'Wyeth/惠氏') & (mydata['sp子品牌'] == '蓝钻') & (mydata['sp产品品牌'] == 'Illuma/启赋'), 'rule'] = 'r8'
    mydata.loc[(mydata['sp厂商'] == 'Wyeth/惠氏') & (mydata['sp子品牌'] == '蕴淳') & (mydata['sp产品品牌'] == 'Illuma/启赋'), 'rule'] = 'r9'
    mydata.loc[(mydata['sp厂商'] == 'Wyeth/惠氏') & (mydata['sp子品牌'] == 'Organic') & (mydata['sp产品品牌'] == 'Illuma/启赋'), 'rule'] = 'r10'
    mydata.loc[(mydata['sp厂商'] == 'Wyeth/惠氏') & (mydata['sp子品牌'] == '铂金') & (mydata['sp产品品牌'] == 'Illuma/启赋'), 'rule'] = 'r11'
    mydata.loc[(mydata['sp厂商'] == 'Mead Johnson/美赞臣') & (mydata['sp产品品牌'] == 'Enfinitas/蓝臻'), 'rule'] = 'r12'

    # 2023-07-04新增sp和p规则
    mydata['sp_rule'] = ''
    mydata.loc[(mydata['sp厂商'] == 'Mead Johnson/美赞臣') & (mydata['sp产品品牌'].isin(['Enfamil A+/安婴儿 A+', 'Enfamil A+', 'Enfamil A2', 'Enfamil AⅡ'])), 'sp_rule'] = 'r1'
    mydata.loc[(mydata['sp厂商'] == 'Mead Johnson/美赞臣') & (mydata['sp产品品牌'] == 'Enfagrow A+/安儿宝 A+'), 'sp_rule'] = 'r2'
    mydata.loc[(mydata['sp厂商'] == 'Mead Johnson/美赞臣') & (mydata['sp产品品牌'] == 'Enfa A+'), 'sp_rule'] = 'r3'
    mydata.loc[(mydata['sp厂商'] == 'Mead Johnson/美赞臣') & (mydata['sp子品牌'] == '卓智') & (mydata['sp产品品牌'] == 'Nutri Power/学优力'), 'sp_rule'] = 'r3'
    mydata.loc[(mydata['sp厂商'] == 'Wyeth/惠氏') & (mydata['sp子品牌'].isin(['Ultima/铂臻', 'Ultima Pdf/铂臻蔼而嘉'])) & (mydata['sp产品品牌'] == 'S-26'), 'sp_rule'] = 'r4'
    mydata.loc[(mydata['sp厂商'] == 'Frieslandcampina/荷兰皇家菲仕兰') & (mydata['sp子品牌'] == '源悦') & (mydata['sp产品品牌'] == 'Friso/美素佳儿'), 'sp_rule'] = 'r5'
    mydata.loc[(mydata['sp厂商'] == 'Firmus/飞鹤') & (mydata['sp产品品牌'] == '超级飞帆'), 'sp_rule'] = 'r6'
    mydata.loc[(mydata['sp厂商'] == 'Firmus/飞鹤') & (mydata['sp产品品牌'] == '舒贝诺'), 'sp_rule'] = 'r7'
    mydata.loc[(mydata['sp厂商'] == 'Firmus/飞鹤') & (mydata['sp产品品牌'] == '星阶优护'), 'sp_rule'] = 'r8'
    mydata.loc[(mydata['sp厂商'] == 'Firmus/飞鹤') & (mydata['sp产品品牌'].isin(['臻爱飞帆', '精粹', '茁然', '飞睿'])), 'sp_rule'] = 'r9'
    mydata.loc[(mydata['sp厂商'] == 'H&Hgroup/健合集团') & (mydata['sp子品牌'] == '贝塔星') & (mydata['sp产品品牌'] == 'Biostime/合生元'), 'sp_rule'] = 'r10'
    mydata.loc[(mydata['sp厂商'] == 'H&Hgroup/健合集团') & (mydata['sp子品牌'].isin(['满乐', '贝素贝加', '素加', 'Terroir/沃蓝', '儿童成长'])) & (mydata['sp产品品牌'] == 'Biostime/合生元'), 'sp_rule'] = 'r11'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'].isin(['帜亲', ''])) & (mydata['sp产品品牌'] == 'Banner Dairy/旗帜'), 'sp_rule'] = 'r12'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'] == '益佳') & (mydata['sp产品品牌'] == 'Banner Dairy/旗帜'), 'sp_rule'] = 'r13'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'] == '诠维爱') & (mydata['sp产品品牌'] == '小小鲁班'), 'sp_rule'] = 'r14'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'] == '诠维爱aii') & (mydata['sp产品品牌'] == '小小鲁班'), 'sp_rule'] = 'r15'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'] == '恬适') & (mydata['sp产品品牌'] == '君乐宝'), 'sp_rule'] = 'r16'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'].isin(['乐臻', '乐星', '淳护', '澳力高'])) & (mydata['sp产品品牌'] == '君乐宝'), 'sp_rule'] = 'r17'
    mydata.loc[(mydata['sp厂商'] == 'Yili/伊利') & (mydata['sp子品牌'] == '菁护') & (mydata['sp产品品牌'] == '金领冠Pro-Kido'), 'sp_rule'] = 'r18'
    mydata.loc[(mydata['sp厂商'] == 'Yili/伊利') & (mydata['sp子品牌'] == '睿护') & (mydata['sp产品品牌'] == '金领冠Pro-Kido'), 'sp_rule'] = 'r19'
    mydata.loc[(mydata['sp厂商'] == 'Yili/伊利') & (mydata['sp子品牌'] == '有机') & (mydata['sp产品品牌'] == '金领冠Pro-Kido'), 'sp_rule'] = 'r20'

    # 2023-07-04新增SP和P规则
    mydata['p_rule'] = ''
    mydata.loc[(mydata['sp厂商'] == 'Mead Johnson/美赞臣') & (mydata['sp产品品牌'] == 'Enfamil A+/铂睿 A+'), 'p_rule'] = 'r1'
    mydata.loc[(mydata['sp厂商'] == 'Mead Johnson/美赞臣') & (mydata['sp产品品牌'].isin(['Enfakid A+/安儿健 A+', 'Enfaschool A+/安学健 A+'])), 'p_rule'] = 'r2'
    mydata.loc[(mydata['sp厂商'] == 'Mead Johnson/美赞臣') & (mydata['sp子品牌'].isin(['', '营护'])) & (mydata['sp产品品牌'] == 'Nutri Power/学优力'), 'p_rule'] = 'r2'
    mydata.loc[(mydata['sp厂商'] == 'Wyeth/惠氏') & (mydata['sp子品牌'] == 'Gold') & (mydata['sp产品品牌'] == 'S-26'), 'p_rule'] = 'r3'
    mydata.loc[(mydata['sp厂商'] == 'Wyeth/惠氏') & (mydata['sp子品牌'].isin(['膳儿加', '学儿乐', '幼儿乐'])) & (mydata['sp产品品牌'] == 'S-26'), 'p_rule'] = 'r4'
    mydata.loc[(mydata['sp厂商'] == 'Frieslandcampina/荷兰皇家菲仕兰') & (mydata['sp子品牌'].isin(['金装', ''])) & (mydata['sp产品品牌'] == 'Friso/美素佳儿'), 'p_rule'] = 'r5'
    mydata.loc[(mydata['sp厂商'] == 'Firmus/飞鹤') & (mydata['sp产品品牌'] == '飞帆'), 'p_rule'] = 'r6'
    mydata.loc[(mydata['sp厂商'] == 'H&Hgroup/健合集团') & (mydata['sp子品牌'] == '阿尔法星') & (mydata['sp产品品牌'] == 'Biostime/合生元'), 'p_rule'] = 'r7'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'].isin(['诠护爱', '诠力爱', ''])) & (mydata['sp产品品牌'] == '小小鲁班'), 'p_rule'] = 'r8'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'].isin(['乐铂', '乐铂k2'])) & (mydata['sp产品品牌'] == '君乐宝'), 'p_rule'] = 'r9'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'] == '至臻') & (mydata['sp产品品牌'] == '君乐宝'), 'p_rule'] = 'r10'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'] == '小旗才') & (mydata['sp产品品牌'] == 'Banner Dairy/旗帜'), 'p_rule'] = 'r11'
    mydata.loc[(mydata['sp厂商'] == 'Junlebao/君乐宝') & (mydata['sp子品牌'].isin(['乐纯', '乐畅'])) & (mydata['sp产品品牌'] == '君乐宝'), 'p_rule'] = 'r11'
    mydata.loc[(mydata['sp厂商'] == 'Yili/伊利') & (mydata['sp子品牌'].isin(['塞纳牧', '赋能', '呵护', '悠滋小羊', '育护', '珍护菁蕴', ''])) & (mydata['sp产品品牌'] == '金领冠Pro-Kido'), 'p_rule'] = 'r12'
    mydata.loc[(mydata['sp厂商'] == 'Yili/伊利') & (mydata['sp子品牌'].isin(['倍冠', '赋能星'])) & (mydata['sp产品品牌'] == '伊利'), 'p_rule'] = 'r13'
    mydata.loc[(mydata['sp厂商'] == 'Yili/伊利') & (mydata['sp产品品牌'] == 'QQ星'), 'p_rule'] = 'r13'
    mydata.loc[(mydata['sp厂商'] == 'Danone/达能') & (mydata['sp子品牌'] != '卓萃') & (mydata['sp产品品牌'] == 'Aptamil/爱他美'), 'p_rule'] = 'r14'
    mydata.loc[(mydata['sp厂商'] == 'Danone/达能') & (mydata['sp产品品牌'].isin(['Nutrilon/诺优能', 'Nutrilon/诺贝能'])), 'p_rule'] = 'r15'

    return mydata

async def fill_total(mydata,table):
    value_type = ['total_sales', 'volume']
    # pannel_type = ['total','up']
    date_type = ['FY21', 'FY22', 'YTD22', 'YTD23']
    people_group = ['月报BaseTotal', '1段', '2段', '3段', '4段']
    table = load_workbook(r'C:\Users\zeng.xiangyan\美素佳儿奶粉月报PPT数据更新\UP&SP&P市场表现_202307_20230821.xlsx')
    sheet_name = {'YTD-UP': ['total', 'up'], 'YTD-SP': ['total', 'sp'], 'YTD-P': ['total', 'p']}
    for sn in sheet_name.keys():
        sheet = table[sn]
        pannel_type = sheet_name[sn]
        s = 0
        for i, v in enumerate(value_type):
            for j, p in enumerate(pannel_type):
                for k, d in enumerate(date_type):
                    for l, g in enumerate(people_group):
                        if g != '月报BaseTotal':
                            if (p == pannel_type[1]) and (d in ['FY21', 'FY22']):
                                data = mydata.loc[
                                    (mydata['是否' + pannel_type[1]] == pannel_type[1]) & (mydata['FY'] == d) & (
                                                mydata['sp适用人群（段数）'] == g) & (mydata['是否限制子品类'] == '限制')]
                            if (p == pannel_type[1]) and (d in ['YTD22', 'YTD23']):
                                data = mydata.loc[
                                    (mydata['是否' + pannel_type[1]] == pannel_type[1]) & (mydata['YTD'] == d) & (
                                                mydata['sp适用人群（段数）'] == g) & (mydata['是否限制子品类'] == '限制')]
                            if (p == 'total') and (d in ['FY21', 'FY22']):
                                data = mydata.loc[(mydata['FY'] == d) & (mydata['sp适用人群（段数）'] == g) & (
                                            mydata['是否限制子品类'] == '限制')]
                            if (p == 'total') and (d in ['YTD22', 'YTD23']):
                                data = mydata.loc[(mydata['YTD'] == d) & (mydata['sp适用人群（段数）'] == g) & (
                                            mydata['是否限制子品类'] == '限制')]
                        else:
                            if (p == pannel_type[1]) and (d in ['FY21', 'FY22']):
                                data = mydata.loc[
                                    (mydata['是否' + pannel_type[1]] == pannel_type[1]) & (mydata['FY'] == d)]
                            if (p == pannel_type[1]) and (d in ['YTD22', 'YTD23']):
                                data = mydata.loc[
                                    (mydata['是否' + pannel_type[1]] == pannel_type[1]) & (mydata['YTD'] == d)]
                            if (p == 'total') and (d in ['FY21', 'FY22']):
                                data = mydata.loc[(mydata['FY'] == d)]
                            if (p == 'total') and (d in ['YTD22', 'YTD23']):
                                data = mydata.loc[(mydata['YTD'] == d)]
                        if d in ['FY21', 'YTD22'] and l == 0:
                            s += 1
                        if (p == 'total') and (d in ['YTD22', 'YTD23']):
                            sheet.cell(8 + l, 2 + s + i * (len(pannel_type) * len(date_type)) + j * (len(date_type)) + k + 3).value = data[v].sum()
                        if (p in ['up', 'sp', 'p']) and (d in ['FY21', 'FY22']) and (g != '月报BaseTotal'):
                            sheet.cell(8 + l, 2 + s + i * (len(pannel_type) * len(date_type)) + j * (len(date_type)) + k - 3).value = data[v].sum()
                        if (p == 'total') and (d in ['FY21', 'FY22']):
                            sheet.cell(8 + l, 2 + s + i * (len(pannel_type) * len(date_type)) + j * (len(date_type)) + k).value = data[v].sum()
                        if (p in ['up', 'sp', 'p']) and (d in ['YTD22', 'YTD23']) and (g != '月报BaseTotal'):
                            sheet.cell(8 + l, 2 + s + i * (len(pannel_type) * len(date_type)) + j * (len(date_type)) + k).value = data[v].sum()
                        # print(8 + l, 2 + s + i * (len(pannel_type) * len(date_type)) + j * (len(date_type)) + k)
        for i in range(8):
            for j in range(len(people_group)):
                # print(8 + j, 5 + i * 3)
                try:
                    sheet.cell(8 + j, 5 + i * 3).value = (sheet.cell(8 + j, 5 + i * 3 - 1).value - sheet.cell(8 + j,5 + i * 3 - 2).value) / sheet.cell(8 + j, 5 + i * 3 - 2).value
                except:
                    continue
    return table

async def fill_detail(mydata,table):
    value_type = ['total_sales', 'volume']
    date_type = ['FY21', 'FY22', 'YTD22', 'YTD23']
    people_group = ['1段', '2段', '3段', '4段']
    rule_dict = {'rule': ['r1', 'r2', 'r3', 'r4', 'r5', 'r6', 'r7', 'r8', 'r9', 'r10', 'r11', 'r12'],
                 'sp_rule': ['r1', 'r2', 'r3', 'r4', 'r5', 'r6', 'r7', 'r8', 'r9', 'r10', 'r11', 'r12', 'r13', 'r14','r15', 'r16', 'r17', 'r18', 'r19', 'r20'],
                 'p_rule': ['r1', 'r2', 'r3', 'r4', 'r5', 'r6', 'r7', 'r8', 'r9', 'r10', 'r11', 'r12', 'r13', 'r14','r15']}
    sheet_name = {'YTD-UP': ['total', 'up'], 'YTD-SP': ['total', 'sp'], 'YTD-P': ['total', 'p']}
    for nr, sn in enumerate(sheet_name.keys()):
        sheet = table[sn]
        pannel_type = [sheet_name[sn][1]]
        rule = rule_dict[list(rule_dict.keys())[nr]]
        s = 0
        for i, v in enumerate(value_type):
            for j, p in enumerate(pannel_type):
                for k, d in enumerate(date_type):
                    for l, g in enumerate(people_group):
                        if d in ['FY21', 'YTD22'] and l == 0:
                            s += 1
                        if (p == sheet_name[sn][1]) and (d in ['FY21', 'FY22']):
                            data = mydata.loc[(mydata['是否' + sheet_name[sn][1]] == sheet_name[sn][1]) & (mydata['FY'] == d) & (mydata['sp适用人群（段数）'] == g) & (mydata['是否限制子品类'] == '限制')]
                        if (p == sheet_name[sn][1]) and (d in ['YTD22', 'YTD23']):
                            data = mydata.loc[(mydata['是否' + sheet_name[sn][1]] == sheet_name[sn][1]) & (mydata['YTD'] == d) & (mydata['sp适用人群（段数）'] == g) & (mydata['是否限制子品类'] == '限制')]
                        sheet.cell(16 + l * (len(rule) + 1),2 + s + i * (len(pannel_type) * len(date_type)) + j * (len(date_type)) + k).value = data[v].sum()

                        for m, r in enumerate(rule):
                            if (p == sheet_name[sn][1]) and (d in ['FY21', 'FY22']):
                                data = mydata.loc[(mydata['是否' + sheet_name[sn][1]] == sheet_name[sn][1]) & (mydata['FY'] == d) & (mydata['sp适用人群（段数）'] == g) & (mydata[list(rule_dict.keys())[nr]] == r) & (mydata['是否限制子品类'] == '限制')]
                            if (p == sheet_name[sn][1]) and (d in ['YTD22', 'YTD23']):
                                data = mydata.loc[(mydata['是否' + sheet_name[sn][1]] == sheet_name[sn][1]) & (mydata['YTD'] == d) & (mydata['sp适用人群（段数）'] == g) & (mydata[list(rule_dict.keys())[nr]] == r) & (mydata['是否限制子品类'] == '限制')]
                            if (p == 'total') and (d in ['FY21', 'FY22']):
                                data = mydata.loc[(mydata['FY'] == d) & (mydata['sp适用人群（段数）'] == g) & (mydata[list(rule_dict.keys())[nr]] == r) & (mydata['是否限制子品类'] == '限制')]
                            if (p == 'total') and (d in ['YTD22', 'YTD23']):
                                data = mydata.loc[(mydata['YTD'] == d) & (mydata['sp适用人群（段数）'] == g) & (mydata[list(rule_dict.keys())[nr]] == r) & (mydata['是否限制子品类'] == '限制')]

                            sheet.cell(17 + l * (len(rule) + 1) + m,2 + s + i * (len(pannel_type) * len(date_type)) + j * (len(date_type)) + k).value = data[v].sum()

        for i in range(4):
            for j in range(len(people_group) * (len(rule) + 1)):
                try:
                    sheet.cell(16 + j, 5 + i * 3).value = (sheet.cell(16 + j, 5 + i * 3 - 1).value - sheet.cell(16 + j,5 + i * 3 - 2).value) / sheet.cell(16 + j, 5 + i * 3 - 2).value
                except:
                    sheet.cell(16 + j, 5 + i * 3).value = '-'
    return table

async def get_newdata(d0,d1,d2,d3,d4,work_book):
    sql_list = sql_date_info(d0,d1,d2,d3,d4)

    for n,sql in enumerate(sql_list):
        print(sql)
        df = await async_connect(0,sql)
        mydata = clean(df)
        table = await fill_total(mydata=mydata,table=work_book)
        table = await fill_detail(mydata=mydata,table=table)

    return table

def run(start_date,end_date):
    try:
        file_path = r'C:/Users/zeng.xiangyan/Desktop/my_sop/my_sop/media/batch172_excel/'
        report_date = end_date
        d0 = str(int(datetime.strptime(report_date, '%Y-%m-%d').strftime('%Y')) - 2) + '-01-01'
        d1 = str(int(datetime.strptime(report_date, '%Y-%m-%d').strftime('%Y'))-1) + '-01-01'
        d2 = (datetime.strptime(report_date, '%Y-%m-%d') - relativedelta(months=12)).strftime('%Y-%m-%d')
        d3 = datetime.strptime(report_date, '%Y-%m-%d').strftime('%Y') + '-01-01'
        d4 = end_date
        print(d0,d1,d2,d3,d4)
        file_name = 'UP&SP&P市场表现_202307_20230821.xlsx'
        work_book = load_workbook(file_path + file_name)
        work_book = asyncio.run(get_newdata(d0,d1,d2,d3,d4, work_book=work_book))
        work_book.save(file_path + f'UP&SP&P市场表现_OUTPUT_{datetime.now().strftime("%Y%m%d")}.xlsx')
        return 1,f'UP%26SP%26P市场表现_OUTPUT_{datetime.now().strftime("%Y%m%d")}.xlsx'
    except Exception as e:
        print(e)
        return 0,'_'
if __name__ == "__main__":
    run('2021-01-01','2023-12-01')