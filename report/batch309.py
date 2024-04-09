# -*- coding: utf-8 -*-
import sys
from os.path import abspath, join, dirname
from datetime import datetime

import pandas as pd
from dateutil.relativedelta import relativedelta
import asyncio
from openpyxl import load_workbook
sys.path.insert(0, join(abspath(dirname(__file__)), '../'))

from sop.connect_clickhouse import async_connect

def sql_date_info(d1,d2,d3,d4):
    sql_list = [f"""
        SELECT * FROM(
            SELECT toStartOfMonth(pkey) AS Period,
            CASE 
                WHEN `sp渠道（新）`in ['TM Flagship','TM B Store','TM Supermarket','TM HK'] THEN 'tmall'
                WHEN `sp渠道（新）`in ['C2C'] THEN 'C2C'
            ELSE 'Others' END as Channel,`sp渠道（新）` as `Sub-channel`,
            CASE 
                WHEN alias_all_bid in (
                    SELECT alias_all_bid FROM(
                        SELECT alias_all_bid,SUM(sales)/100 as `销售额` FROM sop_e.entity_prod_92162_E_0523
                        WHERE ((`date`>='{d3}' AND `date`<'{d4}')
                        AND (`sp子品类` in ['女刀'])
                        AND (source*100+shop_type in (109,121,122,123,124,125,126,127,128)))
                        GROUP by alias_all_bid
                        ORDER by `销售额` DESC
                        LIMIT 5)) THEN dictGetOrDefault('all_brand', 'name', tuple(toUInt32(alias_all_bid)), '')
                ELSE 'others' END as Brand,SUM(sales)/100 as Value,SUM(num) as Volume,SUM(sales)/(100*SUM(num)) as AUS
            FROM sop_e.entity_prod_92162_E_0523
            WHERE ((`date`>='{d1}' AND `date`<'{d2}')
            AND (`sp子品类` in ['女刀'])
            AND (`source` = 1))
            GROUP BY Period,Channel,`Sub-channel`,Brand
            ORDER by Value DESC)
        WHERE Channel != 'C2C';""",
            f"""
        SELECT toStartOfMonth(pkey) AS Period,
        CASE 
            WHEN `sp渠道（新）`in ['TM Flagship','TM B Store','TM Supermarket','TM HK'] THEN 'tmall'
            WHEN `sp渠道（新）`in ['C2C'] THEN 'C2C'
        ELSE 'Others' END as Channel,`sp渠道（新）` as `Sub-channel`,
        CASE 
            WHEN alias_all_bid in (
                SELECT alias_all_bid FROM(
                    SELECT alias_all_bid,SUM(sales)/100 as `销售额` FROM sop_e.entity_prod_92162_E_0523
                    WHERE ((`date`>='{d3}' AND `date`<'{d4}')
                        AND (`sp子品类（女刀月报）` in ['女刀'])
                        AND (source*100+shop_type in (109,121,122,123,124,125,126,127,128)))
                    GROUP by alias_all_bid
                    ORDER by `销售额` DESC
                    LIMIT 5)) THEN dictGetOrDefault('all_brand', 'name', tuple(toUInt32(alias_all_bid)), '')
            ELSE 'others' END as Brand,SUM(sales)/100 as Value,SUM(num) as Volume,SUM(sales)/(100*SUM(num)) as AUS
        FROM sop_e.entity_prod_92162_E_0523
        WHERE ((`date`>='{d3}' AND `date`<'{d4}')
            AND (`sp子品类（女刀月报）` in ['女刀'])
            AND (`source` = 1))
        GROUP BY Period,Channel,`Sub-channel`,Brand
        ORDER by Value DESC;"""]
    return sql_list

async def write_newdata(df,work_book,start_row):

    sheet = work_book['Schick女刀月度销售数据_情报通']
    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            sheet.cell(start_row+r,c+1).value = df.iloc[r].to_list()[c]
    return work_book

async def get_newdata(start_date,end_date,work_book,start_row):
    d3,d4 = start_date,end_date
    d1 = (datetime.strptime(start_date, '%Y-%m-%d') - relativedelta(months=12)).strftime('%Y-%m-%d')
    d2 = (datetime.strptime(end_date, '%Y-%m-%d') - relativedelta(months=12)).strftime('%Y-%m-%d')
    sql_list = sql_date_info(d1,d2,d3,d4)

    count = []
    for n,sql in enumerate(sql_list):
        print(sql)
        df = await async_connect(0,sql)
        count.append(df.shape[0])
        work_book = await write_newdata(df,work_book,start_row + n*count[0])

    return work_book

def run(start_date,end_date):

    try:
        file_path = r'C:/Users/zeng.xiangyan/Desktop/my_sop/my_sop/media/batch309/'
        report_date = end_date
        tt1 = (datetime.strptime(report_date, '%Y-%m-%d') - relativedelta(months=2)).strftime('%Y%m')
        tt2 = (datetime.strptime(report_date, '%Y-%m-%d') - relativedelta(months=1)).strftime('%Y%m')
        file_name = f'【{tt1}】Schick女刀月度销售数据_情报通.xlsx'
        start_row = pd.read_excel(file_path + file_name).shape[0] + 2
        work_book = load_workbook(file_path + file_name)
        work_book = asyncio.run(get_newdata(start_date,end_date, work_book=work_book, start_row=start_row))
        work_book.save(file_path + f'【{tt2}】Schick女刀月度销售数据_情报通.xlsx')
        return 1,f'【{tt2}】Schick女刀月度销售数据_情报通.xlsx'
    except Exception as e:
        print(e)
        return 0,'_'
# if __name__ == '__main__':
#     run('2023-10-01','2023-11-01')